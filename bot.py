# -*- coding: utf-8 -*-

import asyncio
import logging
import sqlite3
import os
from pathlib import Path
from datetime import datetime, timedelta
import random
import string
from io import BytesIO

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, FSInputFile, BufferedInputFile
)
from aiogram.utils.keyboard import InlineKeyboardBuilder

# ===================== ПОЛИТИКА КОНФИДЕНЦИАЛЬНОСТИ =====================
POLICY_URL = "https://telegra.ph/POLITIKA-KONFIDENCIALNOSTI-08-27-64"

# ===================== ЗАГРУЗКА .ENV =====================
try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

# ===================== ДЛЯ ЭКСПОРТА В EXCEL =====================
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("⚠️ openpyxl не установлен. Экспорт в Excel будет недоступен.")

# ===================== ГРЕЧЕСКИЙ АЛФАВИТ =====================
GREEK_LETTERS = [
    "ALPHA", "BETA", "GAMMA", "DELTA", "EPSILON", "ZETA", "ETA", "THETA",
    "IOTA", "KAPPA", "LAMBDA", "MU", "NU", "XI", "OMICRON", "PI",
    "RHO", "SIGMA", "TAU", "UPSILON", "PHI", "CHI", "PSI", "OMEGA"
]

# ===================== ПУТИ =====================
BASE_DIR = Path(__file__).resolve().parent
LOGO_PATH = BASE_DIR / "logo.jpg"
EXAMPLES_DIR = BASE_DIR / "examples"

# ===================== НАСТРОЙКА БАЗЫ ДАННЫХ =====================
DATA_DIR = "/persistent" if os.path.exists("/persistent") else str(BASE_DIR / "data")
os.makedirs(DATA_DIR, exist_ok=True)
DB_NAME = f"{DATA_DIR}/shop_bot.db"

# ===================== НАСТРОЙКИ =====================
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN не задан!")

# ===================== АДМИНИСТРАТОРЫ В КОДЕ =====================
MAIN_ADMINS = []  # Сюда вписывайте ID администраторов (только числа)

# Загружаем админов из .env (для обратной совместимости)
_admins_raw = os.getenv("ADMINS", "")
if _admins_raw:
    for x in _admins_raw.split(","):
        if x.strip().isdigit():
            MAIN_ADMINS.append(int(x.strip()))

ADMINS = MAIN_ADMINS.copy()

if not ADMINS:
    logging.warning("⚠️ Список администраторов пуст! Добавьте ID в MAIN_ADMINS или .env")

DISPATCHER_USERNAME = os.getenv("DISPATCHER_USERNAME", "@sopranidis_support")
CEO_USERNAME = os.getenv("CEO_USERNAME", "@sopranidi")
CHANNEL_LINK = os.getenv("CHANNEL_LINK", "https://t.me/sopranidi_corporation")
BOT_LINK = os.getenv("BOT_LINK", "https://t.me/sopranidi_bot")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


# ===================== АСИНХРОННАЯ ОБЁРТКА =====================
async def run_db(func, *args, **kwargs):
    return await asyncio.to_thread(func, *args, **kwargs)


# ===================== ФУНКЦИИ ГЕНЕРАЦИИ =====================
def generate_order_code() -> str:
    return f"{random.choice(GREEK_LETTERS)}{random.randint(1000, 9999)}"


def generate_unique_order_code() -> str:
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    for _ in range(100):
        code = generate_order_code()
        cur.execute("SELECT order_id FROM orders WHERE order_code = ?", (code,))
        if not cur.fetchone():
            conn.close()
            return code
    timestamp = str(int(datetime.now().timestamp()))[-6:]
    conn.close()
    return f"OMEGA{timestamp}"


def generate_promo_code() -> str:
    return ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(8))


def escape_html(text: str) -> str:
    if not text:
        return ""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def format_price_with_discount(original_price: int, discount: int) -> str:
    if discount <= 0:
        return f"<b>{original_price} ₽</b>"
    new_price = max(1, round(original_price * (1 - discount / 100)))
    return f"<s>{original_price} ₽</s> → <b>{new_price} ₽</b> 🎉"


# ===================== БАЗА ДАННЫХ =====================
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    # Таблица пользователей
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            first_name TEXT,
            last_name TEXT,
            reg_date TEXT,
            last_action TEXT,
            action_date TEXT,
            birthday TEXT DEFAULT '',
            used_promocodes TEXT DEFAULT ''
        )
    """)

    cur.execute("PRAGMA table_info(users)")
    columns = [col[1] for col in cur.fetchall()]

    if "birthday" not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN birthday TEXT DEFAULT ''")
    if "used_promocodes" not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN used_promocodes TEXT DEFAULT ''")
    if "last_action" not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN last_action TEXT DEFAULT ''")
    if "action_date" not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN action_date TEXT DEFAULT ''")
    if "pending_discount" not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN pending_discount INTEGER DEFAULT 0")
    if "pending_discount_code" not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN pending_discount_code TEXT DEFAULT ''")
    if "last_birthday_greet_year" not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN last_birthday_greet_year TEXT DEFAULT ''")

    # Таблица заказов
    cur.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            order_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            service TEXT,
            price INTEGER,
            status TEXT,
            created_at TEXT,
            paid_at TEXT,
            admin_price INTEGER DEFAULT 0,
            admin_note TEXT DEFAULT '',
            order_code TEXT UNIQUE,
            rating INTEGER DEFAULT 0,
            review TEXT DEFAULT '',
            file_id TEXT DEFAULT '',
            is_urgent INTEGER DEFAULT 0,
            discount_applied INTEGER DEFAULT 0
        )
    """)

    # Таблица услуг
    cur.execute("""
        CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            description TEXT,
            price INTEGER,
            is_active INTEGER DEFAULT 1,
            created_at TEXT
        )
    """)

    cur.execute("SELECT COUNT(*) FROM services")
    if cur.fetchone()[0] == 0:
        default_services = [
            ("Курсовая работа", "Помощь в написании курсовой работы по любой теме", 2490),
            ("Школьный проект", "Создание уникального проекта для школы", 1490),
            ("Отчёт по практике", "Оформление отчёта по производственной практике", 2990),
            ("Доклад", "Подготовка качественного доклада на любую тему", 500),
            ("Презентация", "Создание стильной и информативной презентации", 299),
            ("Защитное слово", "Составление защитного слова для проекта", 99),
            ("Реферат", "Написание качественного реферата по любой теме", 1200),
            ("Редактирование работы", "Правка и доработка готовой работы", 800),
            ("Тотальная защита (PREMIER)",
             "Полное сопровождение проекта: консультация, создание работы, объяснение материала, тренаж защиты и финальные правки.",
             3500),
        ]
        for name, desc, price in default_services:
            cur.execute(
                "INSERT INTO services (name, description, price, created_at) VALUES (?, ?, ?, ?)",
                (name, desc, price, datetime.now().isoformat())
            )

    # ===== НОВЫЕ ТАБЛИЦЫ =====

    # Таблица администраторов (для совместимости)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER UNIQUE,
            username TEXT,
            first_name TEXT,
            role TEXT DEFAULT 'admin',
            added_by INTEGER,
            added_at TEXT,
            is_active INTEGER DEFAULT 1
        )
    """)

    # Добавляем супер-админа из MAIN_ADMINS в БД
    if ADMINS:
        for admin_id in ADMINS:
            cur.execute("""
                INSERT OR IGNORE INTO admins (user_id, username, first_name, role, added_at)
                VALUES (?, ?, ?, 'super_admin', ?)
            """, (admin_id, None, 'Главный админ', datetime.now().isoformat()))
        logging.info(f"✅ Супер-админы из MAIN_ADMINS добавлены в БД")

    # Таблица сообщений чата
    cur.execute("""
        CREATE TABLE IF NOT EXISTS chat_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            admin_id INTEGER,
            message TEXT,
            is_from_admin INTEGER DEFAULT 0,
            is_read INTEGER DEFAULT 0,
            created_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(user_id)
        )
    """)

    # Таблица активных чатов
    cur.execute("""
        CREATE TABLE IF NOT EXISTS active_chats (
            user_id INTEGER PRIMARY KEY,
            last_message_at TEXT,
            is_closed INTEGER DEFAULT 0,
            closed_by INTEGER
        )
    """)

    # Голосования
    cur.execute("""
        CREATE TABLE IF NOT EXISTS polls (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            question TEXT,
            options TEXT,
            created_by INTEGER,
            created_at TEXT,
            expires_at TEXT,
            is_active INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS poll_votes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            poll_id INTEGER,
            user_id INTEGER,
            option_text TEXT,
            voted_at TEXT,
            FOREIGN KEY(poll_id) REFERENCES polls(id),
            FOREIGN KEY(user_id) REFERENCES users(user_id)
        )
    """)

    # Промокоды
    cur.execute("""
        CREATE TABLE IF NOT EXISTS promocodes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE,
            discount INTEGER,
            valid_until TEXT,
            max_uses INTEGER,
            used INTEGER DEFAULT 0,
            created_by INTEGER,
            created_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_promocodes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            promo_id INTEGER,
            used_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(user_id),
            FOREIGN KEY(promo_id) REFERENCES promocodes(id)
        )
    """)

    # Акции
    cur.execute("""
        CREATE TABLE IF NOT EXISTS promotions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            description TEXT,
            discount INTEGER,
            valid_until TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT,
            service_id INTEGER DEFAULT 0
        )
    """)

    # Логи
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_logs (
            log_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT,
            details TEXT,
            timestamp TEXT,
            FOREIGN KEY(user_id) REFERENCES users(user_id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS admin_logs (
            log_id INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id INTEGER,
            action TEXT,
            details TEXT,
            timestamp TEXT
        )
    """)

    conn.commit()
    conn.close()
    logging.info("✅ База данных проверена/создана!")


# ===================== ФУНКЦИИ РАБОТЫ С БАЗОЙ =====================

def get_all_admins():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, username, first_name, role, added_at
        FROM admins WHERE is_active = 1 ORDER BY role DESC, added_at ASC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows


def get_admin(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT user_id, username, first_name, role, added_at FROM admins WHERE user_id = ? AND is_active = 1",
                (user_id,))
    row = cur.fetchone()
    conn.close()
    return row


def add_admin(user_id: int, username: str = None, first_name: str = None, added_by: int = None):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    cur.execute("SELECT user_id FROM users WHERE user_id = ?", (user_id,))
    if not cur.fetchone():
        cur.execute(
            "INSERT OR IGNORE INTO users (user_id, username, first_name, reg_date) VALUES (?, ?, ?, ?)",
            (user_id, username, first_name, datetime.now().isoformat())
        )

    cur.execute("""
        INSERT OR IGNORE INTO admins (user_id, username, first_name, role, added_by, added_at)
        VALUES (?, ?, ?, 'admin', ?, ?)
    """, (user_id, username, first_name, added_by, datetime.now().isoformat()))
    conn.commit()
    conn.close()

    if user_id not in MAIN_ADMINS:
        MAIN_ADMINS.append(user_id)
        global ADMINS
        ADMINS = MAIN_ADMINS.copy()


def remove_admin(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE admins SET is_active = 0 WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

    if user_id in MAIN_ADMINS:
        MAIN_ADMINS.remove(user_id)
        global ADMINS
        ADMINS = MAIN_ADMINS.copy()


def is_admin_db(user_id: int) -> bool:
    return user_id in MAIN_ADMINS


def is_super_admin(user_id: int) -> bool:
    return user_id in MAIN_ADMINS


def get_admin_role(user_id: int) -> str:
    if user_id in MAIN_ADMINS:
        return 'super_admin'
    return None


def send_message(user_id: int, admin_id: int, message: str, is_from_admin: int = 0):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO chat_messages (user_id, admin_id, message, is_from_admin, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (user_id, admin_id, message, is_from_admin, datetime.now().isoformat()))

    cur.execute("""
        INSERT OR REPLACE INTO active_chats (user_id, last_message_at)
        VALUES (?, ?)
    """, (user_id, datetime.now().isoformat()))

    conn.commit()
    conn.close()


def get_chat_history(user_id: int, limit: int = 50):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, message, is_from_admin, created_at
        FROM chat_messages
        WHERE user_id = ?
        ORDER BY created_at ASC
        LIMIT ?
    """, (user_id, limit))
    rows = cur.fetchall()
    conn.close()
    return rows


def get_unread_count(user_id: int = None):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    if user_id:
        cur.execute("""
            SELECT COUNT(*) FROM chat_messages
            WHERE user_id = ? AND is_from_admin = 0 AND is_read = 0
        """, (user_id,))
    else:
        cur.execute("""
            SELECT COUNT(*) FROM chat_messages
            WHERE is_from_admin = 0 AND is_read = 0
        """)
    count = cur.fetchone()[0]
    conn.close()
    return count


def mark_as_read(user_id: int, admin_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        UPDATE chat_messages
        SET is_read = 1
        WHERE user_id = ? AND is_from_admin = 0 AND is_read = 0
    """, (user_id,))
    conn.commit()
    conn.close()


def get_active_chats():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT ac.user_id, u.username, u.first_name, ac.last_message_at,
               (SELECT COUNT(*) FROM chat_messages WHERE user_id = ac.user_id AND is_from_admin = 0 AND is_read = 0) as unread
        FROM active_chats ac
        JOIN users u ON ac.user_id = u.user_id
        WHERE ac.is_closed = 0
        ORDER BY ac.last_message_at DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows


def close_chat(user_id: int, admin_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        UPDATE active_chats
        SET is_closed = 1, closed_by = ?
        WHERE user_id = ?
    """, (admin_id, user_id))
    conn.commit()
    conn.close()


def get_user_by_id(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT user_id, username, first_name, last_name FROM users WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    return row


def is_admin(user_id: int) -> bool:
    return is_admin_db(user_id)


def add_user(user_id: int, username: str, first_name: str, last_name: str = ""):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO users (user_id, username, first_name, last_name, reg_date) VALUES (?, ?, ?, ?, ?)",
        (user_id, username, first_name, last_name, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


def update_user_action(user_id: int, action: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET last_action = ?, action_date = ? WHERE user_id = ?",
        (action, datetime.now().isoformat(), user_id)
    )
    conn.commit()
    conn.close()


def add_user_log(user_id: int, action: str, details: str = ""):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO user_logs (user_id, action, details, timestamp) VALUES (?, ?, ?, ?)",
        (user_id, action, details, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


def add_admin_log(admin_id: int, action: str, details: str = ""):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO admin_logs (admin_id, action, details, timestamp) VALUES (?, ?, ?, ?)",
        (admin_id, action, details, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


def get_user(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT user_id, username, first_name, last_name, reg_date, birthday, used_promocodes FROM users WHERE user_id = ?",
        (user_id,)
    )
    row = cur.fetchone()
    conn.close()
    return row


def set_user_birthday(user_id: int, birthday: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE users SET birthday = ? WHERE user_id = ?", (birthday, user_id))
    conn.commit()
    conn.close()


def add_used_promocode(user_id: int, promo_code: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT used_promocodes FROM users WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    used = row[0] if row and row[0] else ""
    new_used = f"{used},{promo_code}" if used else promo_code
    cur.execute("UPDATE users SET used_promocodes = ? WHERE user_id = ?", (new_used, user_id))
    conn.commit()
    conn.close()


def set_pending_discount(user_id: int, discount: int, code: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET pending_discount = ?, pending_discount_code = ? WHERE user_id = ?",
        (discount, code, user_id)
    )
    conn.commit()
    conn.close()


def get_pending_discount(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT pending_discount, pending_discount_code FROM users WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return 0, ""
    return row[0] or 0, row[1] or ""


def clear_pending_discount(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE users SET pending_discount = 0, pending_discount_code = '' WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()


def get_users_with_birthday_today(today_str: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT user_id, first_name, last_birthday_greet_year FROM users WHERE birthday = ?",
        (today_str,)
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def mark_birthday_greeted(user_id: int, year: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE users SET last_birthday_greet_year = ? WHERE user_id = ?", (year, user_id))
    conn.commit()
    conn.close()


def add_order(user_id: int, service: str, price: int, is_urgent: int = 0, discount_applied: int = 0) -> tuple:
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    order_code = generate_unique_order_code()
    cur.execute(
        "INSERT INTO orders (user_id, service, price, status, created_at, order_code, is_urgent, discount_applied) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, service, price, "pending", datetime.now().isoformat(), order_code, is_urgent, discount_applied)
    )
    order_id = cur.lastrowid
    conn.commit()
    conn.close()
    return order_id, order_code


def update_order_status(order_id: int, status: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "UPDATE orders SET status = ?, paid_at = ? WHERE order_id = ?",
        (status, datetime.now().isoformat(), order_id)
    )
    conn.commit()
    conn.close()


def update_order_price(order_id: int, admin_price: int, admin_note: str = ""):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "UPDATE orders SET admin_price = ?, admin_note = ? WHERE order_id = ?",
        (admin_price, admin_note, order_id)
    )
    conn.commit()
    conn.close()


def update_order_review(order_id: int, rating: int, review: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "UPDATE orders SET rating = ?, review = ? WHERE order_id = ?",
        (rating, review, order_id)
    )
    conn.commit()
    conn.close()


def update_order_file(order_id: int, file_id: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "UPDATE orders SET file_id = ? WHERE order_id = ?",
        (file_id, order_id)
    )
    conn.commit()
    conn.close()


def delete_order(order_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM orders WHERE order_id = ?", (order_id,))
    conn.commit()
    conn.close()


def delete_old_orders(days: int = 30):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cutoff_date = (datetime.now() - timedelta(days=days)).isoformat()
    cur.execute(
        "DELETE FROM orders WHERE created_at < ? AND status IN ('paid', 'cancelled')",
        (cutoff_date,)
    )
    deleted_count = cur.rowcount
    conn.commit()
    conn.close()
    return deleted_count


def get_order(order_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT order_id, user_id, service, price, status, created_at, paid_at, admin_price, admin_note, order_code, rating, review, file_id, is_urgent
        FROM orders WHERE order_id = ?
    """, (order_id,))
    row = cur.fetchone()
    conn.close()
    return row


def get_user_orders(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT order_id, service, price, status, created_at, admin_price, order_code, rating, review, is_urgent FROM orders WHERE user_id = ? ORDER BY created_at DESC",
        (user_id,)
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_all_orders():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT o.order_id, o.user_id, u.username, o.service, o.price, o.status, o.created_at, o.paid_at, o.admin_price, o.admin_note, o.order_code, o.rating, o.review, o.is_urgent
        FROM orders o
        LEFT JOIN users u ON o.user_id = u.user_id
        ORDER BY o.created_at DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows


def get_user_stats(user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM orders WHERE user_id = ?", (user_id,))
    total_orders = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM orders WHERE user_id = ? AND status='paid'", (user_id,))
    paid_orders = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM orders WHERE user_id = ? AND status='in_progress'", (user_id,))
    in_progress = cur.fetchone()[0]
    cur.execute("SELECT SUM(price) FROM orders WHERE user_id = ? AND status='paid'", (user_id,))
    total_spent = cur.fetchone()[0] or 0
    conn.close()
    return {
        "total_orders": total_orders,
        "paid_orders": paid_orders,
        "in_progress": in_progress,
        "total_spent": total_spent
    }


def get_all_users():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT user_id, username, first_name, last_name, reg_date, last_action, action_date, birthday FROM users ORDER BY reg_date DESC")
    rows = cur.fetchall()
    conn.close()
    return rows


def get_all_reviews():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT o.order_code, o.service, o.rating, o.review, u.username, u.first_name, o.created_at
        FROM orders o
        LEFT JOIN users u ON o.user_id = u.user_id
        WHERE o.rating > 0 AND o.review != ''
        ORDER BY o.created_at DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows


def get_user_logs(user_id: int, limit: int = 20):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT action, details, timestamp FROM user_logs WHERE user_id = ? ORDER BY timestamp DESC LIMIT ?",
        (user_id, limit)
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_stats():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    user_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM orders")
    total_orders = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM orders WHERE status='paid'")
    paid_orders = cur.fetchone()[0]
    cur.execute("SELECT SUM(price) FROM orders WHERE status='paid'")
    total_income = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM orders WHERE status='pending'")
    pending_orders = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM orders WHERE status='cancelled'")
    cancelled_orders = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM orders WHERE status='in_progress'")
    in_progress = cur.fetchone()[0]
    cur.execute("SELECT AVG(rating) FROM orders WHERE rating > 0")
    avg_rating = cur.fetchone()[0] or 0
    conn.close()
    return {
        "users": user_count,
        "total_orders": total_orders,
        "paid_orders": paid_orders,
        "pending_orders": pending_orders,
        "cancelled_orders": cancelled_orders,
        "in_progress": in_progress,
        "income": total_income,
        "avg_rating": round(avg_rating, 1)
    }


# ===================== УПРАВЛЕНИЕ УСЛУГАМИ =====================
def get_all_services():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT id, name, description, price, is_active FROM services ORDER BY id")
    rows = cur.fetchall()
    conn.close()
    return rows


def get_service(service_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT id, name, description, price, is_active FROM services WHERE id = ?", (service_id,))
    row = cur.fetchone()
    conn.close()
    return row


def add_service(name: str, description: str, price: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO services (name, description, price, created_at) VALUES (?, ?, ?, ?)",
        (name, description, price, datetime.now().isoformat())
    )
    service_id = cur.lastrowid
    conn.commit()
    conn.close()
    return service_id


def update_service(service_id: int, name: str, description: str, price: int, is_active: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "UPDATE services SET name = ?, description = ?, price = ?, is_active = ? WHERE id = ?",
        (name, description, price, is_active, service_id)
    )
    conn.commit()
    conn.close()


def delete_service(service_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM services WHERE id = ?", (service_id,))
    conn.commit()
    conn.close()


def get_services_keyboard(services: list) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    for service in services:
        s_id, name, desc, price, is_active = service
        status = "✅" if is_active else "❌"
        builder.button(text=f"{status} {name} ({price}₽)", callback_data=f"service_edit_{s_id}")
    builder.button(text="➕ Добавить услугу", callback_data="service_add")
    builder.button(text="🔙 Назад", callback_data="admin_menu")
    builder.adjust(1)
    return builder.as_markup()


# ===================== УПРАВЛЕНИЕ АКЦИЯМИ =====================
def create_promotion(name: str, description: str, discount: int, valid_until: str, service_id: int = 0):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO promotions (name, description, discount, valid_until, service_id, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (name, description, discount, valid_until, service_id, datetime.now().isoformat())
    )
    promo_id = cur.lastrowid
    conn.commit()
    conn.close()
    return promo_id


def get_active_promotions(service_id: int = 0):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    if service_id > 0:
        cur.execute(
            "SELECT id, name, description, discount, valid_until, service_id FROM promotions "
            "WHERE is_active = 1 AND valid_until > ? AND (service_id = ? OR service_id = 0) "
            "ORDER BY discount DESC",
            (datetime.now().isoformat(), service_id)
        )
    else:
        cur.execute(
            "SELECT id, name, description, discount, valid_until, service_id FROM promotions "
            "WHERE is_active = 1 AND valid_until > ? ORDER BY discount DESC",
            (datetime.now().isoformat(),)
        )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_all_promotions():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, name, description, discount, valid_until, is_active, created_at, service_id FROM promotions ORDER BY id DESC"
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_promotion_by_id(promo_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, name, description, discount, valid_until, is_active, service_id FROM promotions WHERE id = ?",
        (promo_id,))
    row = cur.fetchone()
    conn.close()
    return row


def activate_promotion(promo_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE promotions SET is_active = 1 WHERE id = ?", (promo_id,))
    conn.commit()
    conn.close()


def deactivate_promotion(promo_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE promotions SET is_active = 0 WHERE id = ?", (promo_id,))
    conn.commit()
    conn.close()


def delete_promotion(promo_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM promotions WHERE id = ?", (promo_id,))
    conn.commit()
    conn.close()


def get_service_discount(service_id: int, user_id: int) -> tuple:
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, name, description, discount, valid_until FROM promotions "
        "WHERE is_active = 1 AND valid_until > ? AND service_id = ? "
        "ORDER BY discount DESC LIMIT 1",
        (datetime.now().isoformat(), service_id)
    )
    promo = cur.fetchone()
    cur.execute(
        "SELECT id, name, description, discount, valid_until FROM promotions "
        "WHERE is_active = 1 AND valid_until > ? AND (service_id = 0 OR service_id IS NULL) "
        "ORDER BY discount DESC LIMIT 1",
        (datetime.now().isoformat(),)
    )
    general_promo = cur.fetchone()
    conn.close()
    promo_discount = 0
    promo_name = ""
    if promo and promo[3] > promo_discount:
        promo_discount = promo[3]
        promo_name = promo[1]
    if general_promo and general_promo[3] > promo_discount:
        promo_discount = general_promo[3]
        promo_name = general_promo[1]
    discount, discount_code = get_pending_discount(user_id)
    if discount > promo_discount:
        return discount, discount_code
    return promo_discount, promo_name


# ===================== ПРОМОКОДЫ =====================
def create_promocode(code: str, discount: int, valid_until: str, max_uses: int, admin_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO promocodes (code, discount, valid_until, max_uses, created_by, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (code, discount, valid_until, max_uses, admin_id, datetime.now().isoformat())
    )
    promo_id = cur.lastrowid
    conn.commit()
    conn.close()
    return promo_id


def get_promocode(code: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT id, code, discount, valid_until, max_uses, used FROM promocodes WHERE code = ?", (code,))
    row = cur.fetchone()
    conn.close()
    return row


def use_promocode(promo_id: int, user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE promocodes SET used = used + 1 WHERE id = ?", (promo_id,))
    cur.execute(
        "INSERT INTO user_promocodes (user_id, promo_id, used_at) VALUES (?, ?, ?)",
        (user_id, promo_id, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


def get_all_promocodes():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT id, code, discount, valid_until, max_uses, used FROM promocodes ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return rows


def delete_promocode(promo_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM promocodes WHERE id = ?", (promo_id,))
    conn.commit()
    conn.close()


# ===================== ГОЛОСОВАНИЯ =====================
def create_poll(question: str, options: list, created_by: int, expires_in_hours: int = 24):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    options_str = "||".join(options)
    expires_at = (datetime.now() + timedelta(hours=expires_in_hours)).isoformat()
    cur.execute(
        "INSERT INTO polls (question, options, created_by, created_at, expires_at) VALUES (?, ?, ?, ?, ?)",
        (question, options_str, created_by, datetime.now().isoformat(), expires_at)
    )
    poll_id = cur.lastrowid
    conn.commit()
    conn.close()
    return poll_id


def get_poll(poll_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT id, question, options, created_by, created_at, expires_at, is_active FROM polls WHERE id = ?",
                (poll_id,))
    row = cur.fetchone()
    conn.close()
    return row


def get_all_polls():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT id, question, created_at, expires_at, is_active FROM polls ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return rows


def vote_poll(poll_id: int, user_id: int, option: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT id FROM poll_votes WHERE poll_id = ? AND user_id = ?", (poll_id, user_id))
    if cur.fetchone():
        conn.close()
        return False
    cur.execute(
        "INSERT INTO poll_votes (poll_id, user_id, option_text, voted_at) VALUES (?, ?, ?, ?)",
        (poll_id, user_id, option, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    return True


def get_poll_results(poll_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT option_text, COUNT(*) FROM poll_votes WHERE poll_id = ? GROUP BY option_text", (poll_id,))
    rows = cur.fetchall()
    conn.close()
    return rows


def get_poll_user_vote(poll_id: int, user_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT option_text FROM poll_votes WHERE poll_id = ? AND user_id = ?", (poll_id, user_id))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None


def close_poll(poll_id: int):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE polls SET is_active = 0 WHERE id = ?", (poll_id,))
    conn.commit()
    conn.close()


def get_expired_active_polls():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, question FROM polls WHERE is_active = 1 AND expires_at < ?",
        (datetime.now().isoformat(),)
    )
    rows = cur.fetchall()
    conn.close()
    return rows


# ===================== ЭКСПОРТ ЗАКАЗОВ В EXCEL =====================
def export_orders_to_excel():
    if not EXCEL_AVAILABLE:
        return None
    orders = get_all_orders()
    if not orders:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"
    headers = ["ID", "Код заказа", "Пользователь", "Услуга", "Цена", "Статус", "Дата создания", "Дата оплаты", "Оценка",
               "Отзыв", "Срочный"]
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    for row_idx, order in enumerate(orders, 2):
        order_id, user_id, username, service, price, status, created_at, paid_at, admin_price, admin_note, order_code, rating, review, is_urgent = order[
            :14]
        final_price = admin_price if admin_price > 0 else price
        status_text = {
            "pending": "Ожидает оплаты",
            "paid": "Оплачен",
            "in_progress": "В работе",
            "cancelled": "Отменён"
        }.get(status, status)
        ws.cell(row=row_idx, column=1, value=order_id)
        ws.cell(row=row_idx, column=2, value=order_code or f"#{order_id}")
        ws.cell(row=row_idx, column=3, value=username or f"ID:{user_id}")
        ws.cell(row=row_idx, column=4, value=service)
        ws.cell(row=row_idx, column=5, value=final_price)
        ws.cell(row=row_idx, column=6, value=status_text)
        ws.cell(row=row_idx, column=7, value=created_at[:10] if created_at else "")
        ws.cell(row=row_idx, column=8, value=paid_at[:10] if paid_at else "")
        ws.cell(row=row_idx, column=9, value=rating if rating > 0 else "")
        ws.cell(row=row_idx, column=10, value=(review[:100] + "..." if len(review) > 100 else review) if review else "")
        ws.cell(row=row_idx, column=11, value="Да" if is_urgent else "Нет")
    for col in range(1, len(headers) + 1):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)
    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===================== КЛАВИАТУРЫ =====================
def main_menu_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.button(text="📚 Заказать работу", callback_data="buy")
    builder.button(text="🎉 Акции", callback_data="promotions")
    builder.button(text="📂 Примеры работ", callback_data="examples")
    builder.button(text="📞 Поддержка", callback_data="support")
    builder.button(text="📋 Мои заказы", callback_data="my_orders")
    builder.button(text="👤 Профиль", callback_data="profile")
    builder.button(text="⭐ Отзывы", callback_data="view_reviews")
    builder.button(text="ℹ️ О нас", callback_data="about")
    builder.adjust(2, 2, 2, 1)
    return builder.as_markup()


def admin_menu_keyboard(user_id: int) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    is_super = is_super_admin(user_id)

    # Базовые функции (доступны всем админам)
    builder.button(text="📊 Статистика", callback_data="admin_stats")
    builder.button(text="👥 Пользователи", callback_data="admin_users")
    builder.button(text="📦 Заказы", callback_data="admin_orders")
    builder.button(text="⭐ Управление отзывами", callback_data="admin_reviews")
    builder.button(text="💬 Чаты с клиентами", callback_data="admin_chats")
    builder.button(text="📋 Логи", callback_data="admin_logs")

    # Функции только для супер-админа
    if is_super:
        builder.button(text="📊 Экспорт заказов", callback_data="admin_export_orders")
        builder.button(text="🛠️ Управление услугами", callback_data="admin_services")
        builder.button(text="🎉 Управление акциями", callback_data="admin_promotions")
        builder.button(text="🎯 Голосования", callback_data="admin_polls")
        builder.button(text="🏷️ Промокоды", callback_data="admin_promocodes")
        builder.button(text="🗑️ Удалить старые заказы", callback_data="admin_delete_old")
        builder.button(text="👑 Управление админами", callback_data="admin_admins")

    builder.button(text="🔙 В главное меню", callback_data="main_menu")
    builder.adjust(2, 2, 2, 1)
    return builder.as_markup()


def services_keyboard_from_db(services: list, user_id: int = None) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    emoji_map = {
        "Курсовая работа": "📝",
        "Школьный проект": "🏫",
        "Отчёт по практике": "📊",
        "Доклад": "🎤",
        "Презентация": "📽️",
        "Защитное слово": "🛡️",
        "Реферат": "📚",
        "Редактирование работы": "✏️",
        "Тотальная защита (PREMIER)": "🛡️",
    }
    for service in services:
        s_id, name, desc, price, is_active = service
        if not is_active:
            continue
        display_price = price
        discount_text = ""
        emoji = emoji_map.get(name, "📋")
        if user_id:
            discount, discount_name = get_service_discount(s_id, user_id)
            if discount > 0:
                new_price = max(1, round(price * (1 - discount / 100)))
                display_price = new_price
                discount_text = f" 🔥-{discount}%"
        builder.button(
            text=f"{emoji} {name} (от {display_price}₽){discount_text}",
            callback_data=f"buyservice_{s_id}"
        )
    builder.button(text="🔙 Назад", callback_data="main_menu")
    builder.adjust(1)
    return builder.as_markup()


def service_edit_keyboard(service_id: int) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.button(text="✏️ Редактировать", callback_data=f"service_edit_form_{service_id}")
    builder.button(text="🔄 Вкл/Выкл", callback_data=f"service_toggle_{service_id}")
    builder.button(text="🗑️ Удалить", callback_data=f"service_delete_{service_id}")
    builder.button(text="🔙 Назад", callback_data="admin_services")
    builder.adjust(1)
    return builder.as_markup()


def back_to_admin_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.button(text="🔙 Назад", callback_data="admin_menu")
    return builder.as_markup()


def back_to_main_keyboard() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.button(text="🔙 В главное меню", callback_data="main_menu")
    return builder.as_markup()


def users_keyboard(users: list, page: int = 0) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    start = page * 10
    end = start + 10
    page_users = users[start:end]
    for user in page_users:
        user_id, username, first_name, _, _, _, _, _ = user
        name = username or first_name or str(user_id)
        builder.button(text=f"👤 {name[:15]}", callback_data=f"user_{user_id}")
    nav_buttons = []
    if page > 0:
        nav_buttons.append(("◀️ Назад", f"users_page_{page - 1}"))
    if end < len(users):
        nav_buttons.append(("Вперед ▶️", f"users_page_{page + 1}"))
    for text, data in nav_buttons:
        builder.button(text=text, callback_data=data)
    builder.button(text="🔙 Назад", callback_data="admin_menu")
    builder.adjust(1)
    return builder.as_markup()


def orders_keyboard(orders: list, page: int = 0) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    start = page * 10
    end = start + 10
    page_orders = orders[start:end]
    for order in page_orders:
        order_id, user_id, username, service, price, status, _, _, admin_price, _, order_code, rating, _, is_urgent = order[
            :14]
        status_emoji = "✅" if status == "paid" else "⏳" if status == "pending" else "🔧" if status == "in_progress" else "❌"
        urgent = "🔥" if is_urgent else ""
        final_price = admin_price if admin_price > 0 else price
        display_code = order_code or f"#{order_id}"
        display_name = username or f"ID:{user_id}"
        rating_str = f"⭐{rating}" if rating > 0 else ""
        builder.button(
            text=f"{status_emoji}{urgent} {display_code} - {display_name} ({final_price}₽) {rating_str}",
            callback_data=f"order_{order_id}"
        )
    nav_buttons = []
    if page > 0:
        nav_buttons.append(("◀️ Назад", f"orders_page_{page - 1}"))
    if end < len(orders):
        nav_buttons.append(("Вперед ▶️", f"orders_page_{page + 1}"))
    for text, data in nav_buttons:
        builder.button(text=text, callback_data=data)
    builder.button(text="🔙 Назад", callback_data="admin_menu")
    builder.adjust(1)
    return builder.as_markup()


def order_detail_keyboard(order_id: int, status: str, is_urgent: int = 0) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    if status == "pending":
        builder.button(text="✅ Подтвердить оплату", callback_data=f"confirm_payment_{order_id}")
        builder.button(text="💰 Назначить цену", callback_data=f"set_price_{order_id}")
        builder.button(text="🔧 В работу", callback_data=f"start_work_{order_id}")
        builder.button(text="❌ Удалить заказ", callback_data=f"delete_order_{order_id}")
        if is_urgent:
            builder.button(text="🔥 Принять срочный заказ", callback_data=f"accept_urgent_{order_id}")
    elif status == "in_progress":
        builder.button(text="✅ Завершить работу", callback_data=f"complete_work_{order_id}")
        builder.button(text="❌ Удалить заказ", callback_data=f"delete_order_{order_id}")
    elif status == "paid":
        builder.button(text="📎 Прикрепить файл", callback_data=f"attach_file_{order_id}")
        builder.button(text="❌ Удалить заказ", callback_data=f"delete_order_{order_id}")
    builder.button(text="🔙 Назад к заказам", callback_data="admin_orders")
    builder.adjust(1)
    return builder.as_markup()


def order_user_keyboard(order_id: int, status: str) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    if status == "pending":
        builder.button(text="❌ Отменить заказ", callback_data=f"cancel_order_{order_id}")
    if status == "paid":
        builder.button(text="⭐ Оставить отзыв", callback_data=f"review_order_{order_id}")
    builder.button(text="🔄 Обновить", callback_data=f"refresh_order_{order_id}")
    builder.button(text="🔙 Назад", callback_data="my_orders")
    builder.adjust(1)
    return builder.as_markup()


def reviews_keyboard(reviews: list, page: int = 0) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    start = page * 10
    end = start + 10
    page_reviews = reviews[start:end]
    for review in page_reviews:
        order_code, service, rating, review_text, username, first_name, created_at = review
        name = username or first_name or "Аноним"
        display_text = f"{order_code} - {service[:10]} ⭐{rating} - {name[:10]}"
        builder.button(text=display_text, callback_data=f"review_detail_{order_code}")
    nav_buttons = []
    if page > 0:
        nav_buttons.append(("◀️ Назад", f"reviews_page_{page - 1}"))
    if end < len(reviews):
        nav_buttons.append(("Вперед ▶️", f"reviews_page_{page + 1}"))
    for text, data in nav_buttons:
        builder.button(text=text, callback_data=data)
    builder.button(text="🔙 Назад", callback_data="admin_menu")
    builder.adjust(1)
    return builder.as_markup()


# ===================== МАШИНЫ СОСТОЯНИЙ =====================
class SupportState(StatesGroup):
    waiting_for_message = State()


class AdminBroadcastState(StatesGroup):
    waiting_for_message = State()


class AdminSetPriceState(StatesGroup):
    waiting_for_price = State()
    waiting_for_note = State()


class AdminDeleteOldState(StatesGroup):
    waiting_for_days = State()


class ReviewState(StatesGroup):
    waiting_for_rating = State()
    waiting_for_review = State()


class AttachFileState(StatesGroup):
    waiting_for_file = State()


class AdminServiceAddState(StatesGroup):
    waiting_for_name = State()
    waiting_for_description = State()
    waiting_for_price = State()


class AdminServiceEditState(StatesGroup):
    waiting_for_name = State()
    waiting_for_description = State()
    waiting_for_price = State()


class AdminPollCreateState(StatesGroup):
    waiting_for_question = State()
    waiting_for_options = State()
    waiting_for_expiry = State()


class AdminPromocodeCreateState(StatesGroup):
    waiting_for_discount = State()
    waiting_for_valid_until = State()
    waiting_for_max_uses = State()


class UserPromocodeState(StatesGroup):
    waiting_for_code = State()


class UserBirthdayState(StatesGroup):
    waiting_for_birthday = State()


class AdminPromotionCreateState(StatesGroup):
    waiting_for_name = State()
    waiting_for_service = State()
    waiting_for_description = State()
    waiting_for_discount = State()
    waiting_for_valid_until = State()


class ChatState(StatesGroup):
    sending = State()


# ===================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====================
async def update_message(callback: CallbackQuery, text: str, reply_markup=None, parse_mode="HTML"):
    try:
        await callback.message.edit_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except Exception as e:
        err = str(e)
        if "message is not modified" in err:
            return
        if "there is no text" in err:
            try:
                await callback.message.delete()
            except Exception:
                pass
            await callback.message.answer(text, reply_markup=reply_markup, parse_mode=parse_mode)
        else:
            logging.error(f"Ошибка обновления сообщения: {e}")
            await callback.message.answer(text, reply_markup=reply_markup, parse_mode=parse_mode)


async def send_safe_message(message: Message, text: str, reply_markup=None, parse_mode="HTML"):
    try:
        await message.answer(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except Exception as e:
        logging.error(f"Ошибка отправки сообщения: {e}")
        await message.answer(text, reply_markup=reply_markup)


# ===================== ОБРАБОТЧИКИ КОМАНД =====================
@dp.message(Command("start"))
async def cmd_start(message: Message):
    user = message.from_user
    await run_db(add_user, user.id, user.username, user.first_name, user.last_name or "")
    await run_db(update_user_action, user.id, "start")
    await run_db(add_user_log, user.id, "start", "Запустил бота")
    name = escape_html(user.first_name or "Клиент")
    text = f"""
<b>🏛️ Sopranidi Corporation</b>

Уважаемый(ая) {name}, приветствуем вас в официальном сервисе академических и деловых проектов.

Мы — команда экспертов с многолетним опытом создания уникальных исследовательских работ, проектов и презентаций высочайшего качества.

<b>🔐 Важная информация:</b>
• Каждый готовый продукт абсолютно индивидуален и создаётся под ваш запрос
• Все работы хранятся в децентрализованном дата-центре, гарантирующем полную конфиденциальность и недоступность для третьих лиц
• Мы соблюдаем все стандарты академической этики и оригинальности

Выберите необходимый сервис в меню ниже 👇
"""
    if LOGO_PATH.exists():
        try:
            photo = FSInputFile(str(LOGO_PATH))
            await message.answer_photo(photo=photo, caption=text, reply_markup=main_menu_keyboard(), parse_mode="HTML")
            return
        except Exception as e:
            logging.warning(f"Не удалось отправить логотип: {e}")
    await send_safe_message(message, text, main_menu_keyboard(), "HTML")


@dp.message(Command("admins"))
async def cmd_admins(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.")
        return

    admins = await run_db(get_all_admins)
    if not admins:
        await message.answer("👑 Список администраторов пуст.")
        return

    text = "👑 <b>Список администраторов:</b>\n\n"
    for admin in admins:
        user_id, username, first_name, role, added_at = admin
        role_icon = "⭐" if role == "super_admin" else "👤"
        name = username or first_name or str(user_id)
        text += f"{role_icon} {name} (ID: {user_id}) — {role}\n"

    await message.answer(text, parse_mode="HTML")


@dp.message(Command("help"))
async def cmd_help(message: Message):
    text = """
<b>📖 Помощь</b>

/start - Главное меню
/buy - Выбор услуги
/examples - Примеры работ
/support - Поддержка
/my_orders - Мои заказы
/profile - Мой профиль
/set_birthday - Указать дату рождения
/policy - Политика конфиденциальности

Для администраторов:
/admin - Админ-панель
"""
    await send_safe_message(message, text, back_to_main_keyboard(), "HTML")


@dp.message(Command("policy"))
async def cmd_policy(message: Message):
    text = f"""
📄 <b>ПОЛИТИКА КОНФИДЕНЦИАЛЬНОСТИ</b>
<b>Sopranidi Corporation</b>

🔒 Мы уважаем вашу приватность и защищаем ваши данные.

<b>📌 Что мы собираем:</b>
• Telegram ID, имя, username
• Данные о заказах и истории обращений
• Технические данные для улучшения сервиса

<b>📌 Как мы используем:</b>
• Для обработки заказов и поддержки
• Для улучшения качества обслуживания
• Для информирования (только с вашего согласия)

<b>📌 Ваши права и команды:</b>

🔹 <b>Знать, какие данные мы храним</b>
Напишите команду <code>/mydata</code> в боте или запросите у администратора

🔹 <b>Исправить неточные данные</b>
Используйте команду <code>/profile</code> для изменения

🔹 <b>Удалить свои данные</b>
Напишите команду <code>/deletedata</code> (данные будут удалены в течение 72 часов)

🔹 <b>Отозвать согласие</b>
Используйте команду <code>/revoke</code> или напишите администратору

🔹 <b>Получить копию данных</b>
Запросите у администратора по команде <code>/exportdata</code>

📌 <b>Контакты:</b>
👤 Диспетчер: @sopranidis_support
🏛️ CEO: @sopranidi
📧 Email: kpanaet@gmail.com

📄 <b>Полная версия Политики:</b>
<a href="{POLICY_URL}">{POLICY_URL}</a>

<i>Дата обновления: 27.08.2026</i>
"""
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="📄 Полная Политика", url=POLICY_URL)
    keyboard.button(text="🔙 Назад", callback_data="main_menu")
    keyboard.adjust(1)

    await message.answer(text, reply_markup=keyboard.as_markup(), parse_mode="HTML")


@dp.message(Command("set_birthday"))
async def cmd_set_birthday(message: Message, state: FSMContext):
    await message.answer("🎂 Укажите вашу дату рождения в формате <b>ДД.ММ</b> (например, 15.05)", parse_mode="HTML")
    await state.set_state(UserBirthdayState.waiting_for_birthday)


@dp.message(UserBirthdayState.waiting_for_birthday)
async def process_birthday(message: Message, state: FSMContext):
    user_id = message.from_user.id
    birthday = message.text.strip()
    try:
        datetime.strptime(birthday, "%d.%m")
        await run_db(set_user_birthday, user_id, birthday)
        await message.answer(
            f"✅ Дата рождения <b>{escape_html(birthday)}</b> сохранена! 🎉\nВ этот день мы обязательно поздравим вас!",
            parse_mode="HTML"
        )
        await state.clear()
    except ValueError:
        await message.answer("❌ Неверный формат. Введите дату в формате <b>ДД.ММ</b> (например, 15.05)",
                             parse_mode="HTML")


@dp.message(Command("admin"))
async def cmd_admin(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ У вас нет доступа.")
        return
    await run_db(add_admin_log, message.from_user.id, "admin_panel", "Открыл админ-панель")
    stats = await run_db(get_stats)
    text = f"""
<b>🔐 Админ-панель Sopranidi Corporation</b>

👥 Пользователей: <b>{stats['users']}</b>
📦 Всего заказов: <b>{stats['total_orders']}</b>
✅ Оплаченных: <b>{stats['paid_orders']}</b>
⏳ Ожидают оплаты: <b>{stats['pending_orders']}</b>
🔧 В работе: <b>{stats['in_progress']}</b>
❌ Отменённых: <b>{stats['cancelled_orders']}</b>
💰 Доход: <b>{stats['income']} руб.</b>
⭐ Средняя оценка: <b>{stats['avg_rating']}</b>

📌 Диспетчер: {DISPATCHER_USERNAME}
👤 CEO: {CEO_USERNAME}

👇 Выберите действие:
"""
    await send_safe_message(message, text, admin_menu_keyboard(message.from_user.id), "HTML")


@dp.message(Command("addadmin"))
async def cmd_addadmin(message: Message):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Только супер-админ может добавлять.", parse_mode="HTML")
        return

    parts = message.text.split()
    if len(parts) < 2:
        await message.answer(
            "👑 <b>Добавление администратора</b>\n\n"
            "Использование:\n"
            "<code>/addadmin 123456789</code> — по ID\n"
            "<code>/addadmin @username</code> — по username\n"
            "<code>/addadmin t.me/username</code> — по ссылке\n\n"
            "ℹ️ Пользователь должен хотя бы раз написать боту команду /start",
            parse_mode="HTML"
        )
        return

    user_input = parts[1].strip()
    user_id = None

    if user_input.isdigit():
        user_id = int(user_input)
    elif user_input.startswith("@"):
        username = user_input.replace("@", "")
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM users WHERE username = ?", (username,))
        row = cur.fetchone()
        conn.close()
        if row:
            user_id = row[0]
    elif "t.me/" in user_input:
        username = user_input.split("t.me/")[-1].replace("/", "").strip()
        username = username.replace("@", "")
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM users WHERE username = ?", (username,))
        row = cur.fetchone()
        conn.close()
        if row:
            user_id = row[0]

    if not user_id:
        await message.answer(
            "❌ Не удалось найти пользователя.\n\n"
            "Убедитесь, что:\n"
            "1. Пользователь написал боту /start\n"
            "2. Вы правильно ввели ID или username\n\n"
            "Попробуйте: /addadmin @username",
            parse_mode="HTML"
        )
        return

    if await run_db(is_admin_db, user_id):
        await message.answer("❌ Этот пользователь уже является админом.", parse_mode="HTML")
        return

    user = await run_db(get_user_by_id, user_id)
    username = user[1] if user else None
    first_name = user[2] if user else None

    await run_db(add_admin, user_id, username, first_name, message.from_user.id)
    await run_db(add_admin_log, message.from_user.id, "add_admin", f"Добавил админа {user_id}")

    await message.answer(
        f"✅ Администратор <b>{username or first_name or str(user_id)}</b> успешно добавлен!\n\n"
        f"📌 Теперь он может использовать /admin",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )


@dp.message(Command("stats"))
async def cmd_stats(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ У вас нет прав.")
        return
    stats = await run_db(get_stats)
    text = f"""
<b>📊 Статистика</b>

👥 Пользователей: {stats['users']}
📦 Заказов: {stats['total_orders']}
✅ Оплачено: {stats['paid_orders']}
⏳ Ожидают: {stats['pending_orders']}
🔧 В работе: {stats['in_progress']}
❌ Отменено: {stats['cancelled_orders']}
💰 Доход: {stats['income']} руб.
⭐ Средняя оценка: {stats['avg_rating']}
"""
    await message.answer(text, parse_mode="HTML")


@dp.message(Command("broadcast"))
async def cmd_broadcast(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Только супер-админ может делать рассылку.", parse_mode="HTML")
        return
    text = message.text.replace("/broadcast", "", 1).strip()
    if not text:
        await message.answer("📢 Рассылка\n\nВведите текст для рассылки всем пользователям:", parse_mode="HTML")
        await state.set_state(AdminBroadcastState.waiting_for_message)
        return
    users = await run_db(get_all_users)
    sent = 0
    failed = 0
    for uid in users:
        try:
            await bot.send_message(uid[0], text, parse_mode="HTML")
            sent += 1
            await asyncio.sleep(0.1)
        except Exception as e:
            failed += 1
            logging.debug(f"Не удалось отправить рассылку {uid[0]}: {e}")
    await run_db(add_admin_log, message.from_user.id, "broadcast",
                 f"Отправил рассылку {sent} пользователям ({failed} неудачно)")
    await message.answer(f"✅ Рассылка выполнена. Отправлено {sent} пользователям, не доставлено {failed}.",
                         parse_mode="HTML")


# ===================== ГЛОБАЛЬНЫЙ ОБРАБОТЧИК СООБЩЕНИЙ =====================
@dp.message()
async def handle_all_messages(message: Message, state: FSMContext):
    user_id = message.from_user.id

    current_state = await state.get_state()
    if current_state is not None:
        return

    if await run_db(is_admin_db, user_id):
        return

    if message.text and message.text.startswith("/"):
        return

    if not message.text:
        await message.answer(
            "📎 Я принимаю только текстовые сообщения.\n"
            "Для отправки файлов используйте кнопку '📎 Прикрепить файл' в админ-панели.",
            parse_mode="HTML"
        )
        return

    await run_db(send_message, user_id, None, message.text, 0)
    await run_db(add_user_log, user_id, "chat_message", f"Отправил сообщение: {message.text[:50]}...")

    admins = await run_db(get_all_admins)
    user = await run_db(get_user, user_id)
    name = user[1] or user[2] or str(user_id)

    for admin in admins:
        try:
            await bot.send_message(
                admin[0],
                f"📩 <b>Новое сообщение от клиента!</b>\n\n"
                f"👤 {escape_html(name)}\n"
                f"🆔 ID: {user_id}\n"
                f"💬 {escape_html(message.text)}\n\n"
                f"🔹 Нажмите <b>💬 Чаты с клиентами</b> в админ-панели, чтобы ответить.",
                parse_mode="HTML"
            )
        except Exception as e:
            logging.warning(f"Не удалось уведомить админа {admin[0]}: {e}")

    await message.answer(
        "✅ Ваше сообщение отправлено администратору. Ожидайте ответа!\n"
        "📌 Вы можете продолжать диалог в этом чате.",
        parse_mode="HTML"
    )


# ===================== АКЦИИ (ДЛЯ ПОЛЬЗОВАТЕЛЕЙ) =====================
@dp.callback_query(F.data == "promotions")
async def cb_promotions(callback: CallbackQuery):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "promotions")
    await run_db(add_user_log, user_id, "promotions", "Просмотрел акции")
    promotions = await run_db(get_active_promotions)
    if not promotions:
        text = f"""
<b>🎉 Акции и спецпредложения</b>

На данный момент активных акций нет.

Подпишитесь на наш канал, чтобы первыми узнавать о новых предложениях!

📢 {CHANNEL_LINK}
"""
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="📢 Перейти в канал", url=CHANNEL_LINK)
        keyboard.button(text="🔙 Назад", callback_data="main_menu")
        keyboard.adjust(1)
        await update_message(callback, text, keyboard.as_markup(), "HTML")
        await callback.answer()
        return
    text = "<b>🎉 Активные акции и спецпредложения</b>\n\n<i>Все акции автоматически применяются при оформлении заказа.</i>\n\n"
    for promo in promotions:
        promo_id, name, description, discount, valid_until, service_id = promo
        valid_date = datetime.fromisoformat(valid_until).strftime("%d.%m.%Y")
        service_text = " (все услуги)" if service_id == 0 else ""
        text += f"""
<b>🏷️ {escape_html(name)}</b>{service_text}
📝 {escape_html(description)}
🎯 Скидка: <b>{discount}%</b>
📅 Действует до: {valid_date}

"""
    text += """
<b>📌 Как получить скидку?</b>
Просто оформите заказ — скидка применится автоматически!
"""
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="🔙 Назад", callback_data="main_menu")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


# ===================== ПРОСМОТР ОТЗЫВОВ =====================
@dp.callback_query(F.data == "view_reviews")
async def cb_view_reviews(callback: CallbackQuery):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "view_reviews")
    await run_db(add_user_log, user_id, "view_reviews", "Просмотрел отзывы")
    reviews = await run_db(get_all_reviews)
    if not reviews:
        text = "⭐ <b>Отзывы</b>\n\nПока нет ни одного отзыва. Будьте первым!"
        await update_message(callback, text, back_to_main_keyboard(), "HTML")
        await callback.answer()
        return
    text = "<b>⭐ Отзывы наших клиентов:</b>\n\n"
    for review in reviews[:10]:
        order_code, service, rating, review_text, username, first_name, created_at = review
        name = username or first_name or "Аноним"
        created = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
        stars = "⭐" * rating + "☆" * (5 - rating)
        text += f"📌 <b>{escape_html(order_code)}</b> - {escape_html(service)}\n"
        text += f"👤 {escape_html(name)}\n"
        text += f"{stars} {rating}/5\n"
        if review_text:
            text += f"📝 \"{escape_html(review_text[:100])}{'...' if len(review_text) > 100 else ''}\"\n"
        text += f"📅 {created}\n\n"
    if len(reviews) > 10:
        text += f"📌 Показано 10 из {len(reviews)} отзывов"
    await update_message(callback, text, back_to_main_keyboard(), "HTML")
    await callback.answer()


# ===================== ПРОФИЛЬ =====================
@dp.callback_query(F.data == "profile")
async def cb_profile(callback: CallbackQuery):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "profile")
    await run_db(add_user_log, user_id, "profile", "Просмотрел профиль")
    user = await run_db(get_user, user_id)
    if not user:
        await update_message(callback, "❌ Профиль не найден.", back_to_main_keyboard(), "HTML")
        return
    _, username, first_name, last_name, reg_date, birthday, used_promocodes = user
    stats = await run_db(get_user_stats, user_id)
    discount, discount_code = await run_db(get_pending_discount, user_id)
    name = f"{first_name} {last_name or ''}".strip() or "Пользователь"
    username_str = f"@{username}" if username else "Не указан"
    reg_date_str = datetime.fromisoformat(reg_date).strftime("%d.%m.%Y")
    birthday_str = birthday or "Не указана"
    used_list = used_promocodes.split(",") if used_promocodes else []
    text = f"""
<b>👤 Ваш профиль</b>

Имя: {escape_html(name)}
Username: {escape_html(username_str)}
📅 Регистрация: {reg_date_str}
🎂 День рождения: {birthday_str}

<b>📊 Статистика:</b>
📦 Всего заказов: {stats['total_orders']}
✅ Оплачено: {stats['paid_orders']}
🔧 В работе: {stats['in_progress']}
💰 Всего потрачено: {stats['total_spent']} руб.

<b>🏷️ Использованные промокоды:</b>
"""
    if used_list:
        text += "\n".join([f"• {escape_html(code)}" for code in used_list if code])
    else:
        text += "• Нет использованных промокодов"
    if discount > 0:
        text += f"\n\n🎉 У вас активна скидка <b>{discount}%</b> (промокод {escape_html(discount_code)}) — она применится к следующему заказу!"

    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="🎂 Установить день рождения", callback_data="set_birthday")
    keyboard.button(text="🎟️ Применить промокод", callback_data="apply_promocode")
    keyboard.button(text="📄 Политика конфиденциальности", url=POLICY_URL)
    keyboard.button(text="🔙 Назад", callback_data="main_menu")
    keyboard.adjust(1)

    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data == "set_birthday")
async def cb_set_birthday(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "🎂 Укажите вашу дату рождения в формате <b>ДД.ММ</b> (например, 15.05)",
        reply_markup=back_to_main_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(UserBirthdayState.waiting_for_birthday)
    await callback.answer()


@dp.callback_query(F.data == "apply_promocode")
async def cb_apply_promocode(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "🎟️ Введите промокод:",
        reply_markup=back_to_main_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(UserPromocodeState.waiting_for_code)
    await callback.answer()


@dp.message(UserPromocodeState.waiting_for_code)
async def process_promocode(message: Message, state: FSMContext):
    user_id = message.from_user.id
    code = message.text.strip().upper()
    promo = await run_db(get_promocode, code)
    if not promo:
        await message.answer("❌ Промокод не найден. Проверьте правильность ввода.", parse_mode="HTML")
        return
    promo_id, promo_code, discount, valid_until, max_uses, used = promo
    if datetime.now() > datetime.fromisoformat(valid_until):
        await message.answer("❌ Срок действия промокода истёк.", parse_mode="HTML")
        return
    if used >= max_uses:
        await message.answer("❌ Промокод уже использован максимальное количество раз.", parse_mode="HTML")
        return
    user = await run_db(get_user, user_id)
    if user and user[6] and code in user[6].split(","):
        await message.answer("❌ Вы уже использовали этот промокод.", parse_mode="HTML")
        return
    await run_db(use_promocode, promo_id, user_id)
    await run_db(add_used_promocode, user_id, code)
    await run_db(set_pending_discount, user_id, discount, code)
    await message.answer(
        f"✅ Промокод <b>{escape_html(code)}</b> применён!\n\n🎉 Скидка: <b>{discount}%</b>\n\nСкидка будет автоматически применена к вашему следующему заказу!",
        parse_mode="HTML"
    )
    await state.clear()


# ===================== О КОМПАНИИ =====================
@dp.callback_query(F.data == "about")
async def cb_about(callback: CallbackQuery):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "about")
    await run_db(add_user_log, user_id, "about", "Открыл информацию о компании")
    text = f"""
<b>🏛️ Sopranidi Corporation — Ваш надёжный партнёр</b>

Sopranidi Corp. — экосистема профессиональных решений для студентов, аспирантов и специалистов. Мы создаём уникальные академические и деловые проекты с 2023 года, заслужив доверие более 500 клиентов.

<b>🔒 Наши стандарты качества:</b>
• Каждый проект разрабатывается индивидуально под ваши требования
• Гарантируем 100% оригинальность и соответствие академическим стандартам
• Все работы шифруются и хранятся в децентрализованных дата-центрах
• Строгое соблюдение сроков и конфиденциальность данных

<b>📊 Технологическая безопасность:</b>
Ваши проекты защищены многоуровневой системой шифрования. Доступ к данным имеют только авторизованные сотрудники.

<b>📌 Контакты:</b>
👤 Диспетчер: {DISPATCHER_USERNAME}
👤 CEO: {CEO_USERNAME}

<b>🔗 Официальные ресурсы:</b>
📢 Канал: {CHANNEL_LINK}
🤖 Бот: {BOT_LINK}
"""
    keyboard = InlineKeyboardBuilder()
    dispatcher_username = DISPATCHER_USERNAME.replace("@", "")
    ceo_username = CEO_USERNAME.replace("@", "")
    keyboard.button(text="👤 Связаться с диспетчером", url=f"https://t.me/{dispatcher_username}")
    keyboard.button(text="👤 Связаться с CEO", url=f"https://t.me/{ceo_username}")
    keyboard.button(text="📢 Перейти в канал", url=CHANNEL_LINK)
    keyboard.button(text="🤖 Перейти в бота", url=BOT_LINK)
    keyboard.button(text="📄 Политика конфиденциальности", url=POLICY_URL)
    keyboard.button(text="🔙 Назад", callback_data="main_menu")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


# ===================== АДМИН-ПАНЕЛЬ =====================
@dp.callback_query(F.data == "admin_stats")
async def cb_admin_stats(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    stats = await run_db(get_stats)
    await run_db(add_admin_log, callback.from_user.id, "view_stats", "Просмотрел статистику")
    text = f"""
<b>📊 Статистика Sopranidi Corp.</b>

👥 Пользователей: {stats['users']}
📦 Всего заказов: {stats['total_orders']}
✅ Оплаченных: {stats['paid_orders']}
⏳ Ожидают оплаты: {stats['pending_orders']}
🔧 В работе: {stats['in_progress']}
❌ Отменённых: {stats['cancelled_orders']}
💰 Доход: {stats['income']} руб.
⭐ Средняя оценка: {stats['avg_rating']}
"""
    await update_message(callback, text, admin_menu_keyboard(callback.from_user.id), "HTML")
    await callback.answer()


# ===================== АДМИН: УПРАВЛЕНИЕ АДМИНАМИ =====================
@dp.callback_query(F.data == "admin_admins")
async def cb_admin_admins(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return

    admins = await run_db(get_all_admins)
    text = "👑 <b>Управление администраторами</b>\n\n"

    if admins:
        for admin in admins:
            user_id, username, first_name, role, added_at = admin
            role_icon = "⭐" if role == "super_admin" else "👤"
            name = username or first_name or str(user_id)
            text += f"{role_icon} {name} (ID: {user_id})\n"
            text += f"   Роль: {role}\n"
            text += f"   Добавлен: {added_at[:10]}\n\n"
    else:
        text += "Нет администраторов.\n"

    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="➕ Добавить админа", callback_data="admin_add")
    if is_super_admin(callback.from_user.id):
        keyboard.button(text="🗑️ Удалить админа", callback_data="admin_remove")
    keyboard.button(text="🔄 Обновить список", callback_data="admin_admins")
    keyboard.button(text="🔙 Назад", callback_data="admin_menu")
    keyboard.adjust(1)

    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data == "admin_add")
async def cb_admin_add(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ может добавлять", show_alert=True)
        return

    text = """
👑 <b>Добавление администратора</b>

Введите ID пользователя или @username:

Примеры:
<code>123456789</code>
<code>@username</code>
<code>t.me/username</code>

ℹ️ Пользователь должен хотя бы раз написать боту /start
"""
    await callback.message.edit_text(text, reply_markup=back_to_admin_keyboard(), parse_mode="HTML")
    await state.set_state("waiting_for_admin_id")
    await callback.answer()


@dp.message(StateFilter("waiting_for_admin_id"))
async def process_admin_add(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return

    text = message.text.strip()
    user_id = None

    if text.isdigit():
        user_id = int(text)
    elif "t.me/" in text:
        username = text.split("t.me/")[-1].replace("/", "").strip()
        username = username.replace("@", "")
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM users WHERE username = ?", (username,))
        row = cur.fetchone()
        conn.close()
        if row:
            user_id = row[0]
    elif text.startswith("@"):
        username = text.replace("@", "")
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM users WHERE username = ?", (username,))
        row = cur.fetchone()
        conn.close()
        if row:
            user_id = row[0]

    if not user_id:
        await message.answer(
            "❌ Не удалось определить ID пользователя.\n\n"
            "Попробуйте:\n"
            "1. Отправить ID (число)\n"
            "2. Отправить ссылку t.me/username\n"
            "3. Отправить @username\n\n"
            "Или попросите пользователя написать боту команду /start.",
            parse_mode="HTML"
        )
        return

    if await run_db(is_admin_db, user_id):
        await message.answer("❌ Этот пользователь уже является админом.", parse_mode="HTML")
        await state.clear()
        return

    user = await run_db(get_user_by_id, user_id)
    username = user[1] if user else None
    first_name = user[2] if user else None

    await run_db(add_admin, user_id, username, first_name, message.from_user.id)
    await run_db(add_admin_log, message.from_user.id, "add_admin", f"Добавил админа {user_id}")

    await message.answer(
        f"✅ Администратор <b>{username or first_name or str(user_id)}</b> успешно добавлен!\n\n"
        f"📌 Теперь он может использовать /admin",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )
    await state.clear()


@dp.callback_query(F.data == "admin_remove")
async def cb_admin_remove(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ может удалять", show_alert=True)
        return

    admins = await run_db(get_all_admins)
    if not admins:
        await callback.message.edit_text(
            "👑 Нет администраторов для удаления.",
            reply_markup=back_to_admin_keyboard(),
            parse_mode="HTML"
        )
        await callback.answer()
        return

    text = "👑 <b>Выберите админа для удаления:</b>\n\n"
    keyboard = InlineKeyboardBuilder()
    for admin in admins:
        user_id, username, first_name, role, _ = admin
        if role == "super_admin":
            continue
        name = username or first_name or str(user_id)
        keyboard.button(text=f"🗑️ {name}", callback_data=f"admin_remove_confirm_{user_id}")

    keyboard.button(text="🔙 Назад", callback_data="admin_admins")
    keyboard.adjust(1)

    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("admin_remove_confirm_"))
async def cb_admin_remove_confirm(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ может удалять", show_alert=True)
        return

    user_id = int(callback.data.split("_")[3])
    if user_id == callback.from_user.id:
        await callback.answer("❌ Нельзя удалить самого себя", show_alert=True)
        return

    admin = await run_db(get_admin, user_id)
    if not admin:
        await callback.answer("❌ Администратор не найден", show_alert=True)
        return

    if admin[3] == "super_admin":
        await callback.answer("❌ Нельзя удалить супер-админа", show_alert=True)
        return

    await run_db(remove_admin, user_id)
    await run_db(add_admin_log, callback.from_user.id, "remove_admin", f"Удалил админа {user_id}")

    await callback.answer("✅ Администратор удалён!", show_alert=True)
    await cb_admin_admins(callback)


# ===================== АДМИН: ЧАТЫ С КЛИЕНТАМИ =====================
@dp.callback_query(F.data == "admin_chats")
async def cb_admin_chats(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return

    chats = await run_db(get_active_chats)
    unread = await run_db(get_unread_count)

    text = f"💬 <b>Активные чаты</b>\n\n"
    text += f"📬 Непрочитанных сообщений: <b>{unread}</b>\n\n"

    if not chats:
        text += "Активных чатов нет."
        await update_message(callback, text, admin_menu_keyboard(callback.from_user.id), "HTML")
        await callback.answer()
        return

    keyboard = InlineKeyboardBuilder()
    for chat in chats:
        user_id, username, first_name, last_message, unread_count = chat
        name = username or first_name or str(user_id)
        unread_icon = "🔴" if unread_count > 0 else "🟢"
        keyboard.button(
            text=f"{unread_icon} {name} ({unread_count} непрочитанных)" if unread_count > 0 else f"{unread_icon} {name}",
            callback_data=f"chat_open_{user_id}"
        )

    keyboard.button(text="🔙 Назад", callback_data="admin_menu")
    keyboard.adjust(1)

    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("chat_open_"))
async def cb_chat_open(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return

    user_id = int(callback.data.split("_")[2])
    user = await run_db(get_user_by_id, user_id)
    name = user[1] or user[2] or str(user_id)

    await run_db(mark_as_read, user_id, callback.from_user.id)

    history = await run_db(get_chat_history, user_id)

    text = f"💬 <b>Чат с {escape_html(name)}</b>\n\n"

    if history:
        for msg_id, msg, is_from_admin, created_at in history[-20:]:
            time = datetime.fromisoformat(created_at).strftime("%H:%M")
            sender = "👤 Админ" if is_from_admin else "👤 Клиент"
            text += f"{sender} [{time}]: {escape_html(msg)}\n"
    else:
        text += "История сообщений пуста.\n"

    text += "\n✏️ Напишите сообщение:"

    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="❌ Закрыть чат", callback_data=f"chat_close_{user_id}")
    keyboard.button(text="🔙 Назад", callback_data="admin_chats")
    keyboard.adjust(1)

    await update_message(callback, text, keyboard.as_markup(), "HTML")

    await state.set_state(ChatState.sending)
    await state.update_data(chat_user_id=user_id)

    await callback.answer()


@dp.message(ChatState.sending)
async def chat_send_message(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа", parse_mode="HTML")
        await state.clear()
        return

    data = await state.get_data()
    user_id = data.get("chat_user_id")

    if not user_id:
        await message.answer("❌ Чат не найден.", parse_mode="HTML")
        await state.clear()
        return

    if message.text and message.text.startswith("/"):
        return

    await run_db(send_message, user_id, message.from_user.id, message.text, 1)
    await run_db(add_admin_log, message.from_user.id, "chat_message",
                 f"Отправил сообщение пользователю {user_id}: {message.text[:50]}...")

    try:
        await bot.send_message(
            user_id,
            f"📩 <b>Сообщение от администратора</b>\n\n{escape_html(message.text)}\n\n✏️ Вы можете ответить в этом чате.",
            parse_mode="HTML"
        )
        await message.answer("✅ Сообщение отправлено!", parse_mode="HTML")
    except Exception as e:
        logging.error(f"Ошибка отправки сообщения пользователю {user_id}: {e}")
        await message.answer("❌ Не удалось отправить сообщение пользователю.", parse_mode="HTML")

    await cb_chat_open(message, state)


@dp.callback_query(F.data.startswith("chat_close_"))
async def cb_chat_close(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return

    user_id = int(callback.data.split("_")[2])
    await run_db(close_chat, user_id, callback.from_user.id)
    await run_db(add_admin_log, callback.from_user.id, "close_chat", f"Закрыл чат с пользователем {user_id}")

    await callback.answer("✅ Чат закрыт!", show_alert=True)
    await cb_admin_chats(callback)


# ===================== АДМИН: УПРАВЛЕНИЕ ПРОМОКОДАМИ =====================
@dp.callback_query(F.data == "admin_promocodes")
async def cb_admin_promocodes(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    promocodes = await run_db(get_all_promocodes)
    await run_db(add_admin_log, callback.from_user.id, "view_promocodes", "Просмотрел список промокодов")
    text = "🏷️ <b>Управление промокодами</b>\n\n"
    if promocodes:
        for promo in promocodes:
            p_id, code, discount, valid_until, max_uses, used = promo
            valid = datetime.fromisoformat(valid_until).strftime("%d.%m.%Y")
            text += f"• <b>{escape_html(code)}</b> - {discount}% (исп. {used}/{max_uses}) до {valid}\n"
    else:
        text += "Нет созданных промокодов."
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="➕ Создать промокод", callback_data="promocode_create")
    keyboard.button(text="🔙 Назад", callback_data="admin_menu")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data == "promocode_create")
async def cb_promocode_create(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await callback.message.edit_text(
        "🏷️ <b>Создание промокода</b>\n\nВведите размер скидки в % (только число, например, 15):",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminPromocodeCreateState.waiting_for_discount)
    await callback.answer()


@dp.message(AdminPromocodeCreateState.waiting_for_discount)
async def process_promo_discount(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        return

    try:
        discount = int(message.text.strip())
        if discount <= 0 or discount > 100:
            await message.answer("❌ Скидка должна быть от 1 до 100%. Попробуйте снова:", parse_mode="HTML")
            return
        await state.update_data(discount=discount)
        await message.answer("📅 Введите дату окончания действия в формате <b>ДД.ММ.ГГГГ</b>\nПример: 31.12.2026",
                             parse_mode="HTML")
        await state.set_state(AdminPromocodeCreateState.waiting_for_valid_until)
    except ValueError:
        await message.answer("❌ Введите корректное число. Попробуйте снова:", parse_mode="HTML")


@dp.message(AdminPromocodeCreateState.waiting_for_valid_until)
async def process_promo_valid_until(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        return

    try:
        valid_until = datetime.strptime(message.text.strip(), "%d.%m.%Y")
        if valid_until < datetime.now():
            await message.answer("❌ Дата должна быть в будущем. Попробуйте снова:", parse_mode="HTML")
            return
        await state.update_data(valid_until=valid_until.isoformat())
        await message.answer("🔢 Введите максимальное количество использований (например, 50):", parse_mode="HTML")
        await state.set_state(AdminPromocodeCreateState.waiting_for_max_uses)
    except ValueError:
        await message.answer("❌ Неверный формат. Используйте <b>ДД.ММ.ГГГГ</b>. Попробуйте снова:", parse_mode="HTML")


@dp.message(AdminPromocodeCreateState.waiting_for_max_uses)
async def process_promo_max_uses(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return

    if message.text and message.text.startswith("/"):
        await state.clear()
        return

    try:
        max_uses = int(message.text.strip())
        if max_uses <= 0:
            await message.answer("❌ Количество использований должно быть положительным. Попробуйте снова:",
                                 parse_mode="HTML")
            return
        data = await state.get_data()
        code = generate_promo_code()
        await run_db(create_promocode, code, data['discount'], data['valid_until'], max_uses, message.from_user.id)
        await run_db(add_admin_log, message.from_user.id, "create_promocode",
                     f"Создал промокод {code} ({data['discount']}%)")
        text = (f"✅ <b>Промокод создан!</b>\n\n"
                f"🏷️ Код: <b>{escape_html(code)}</b>\n"
                f"🎉 Скидка: <b>{data['discount']}%</b>\n"
                f"📅 Действует до: {datetime.fromisoformat(data['valid_until']).strftime('%d.%m.%Y')}\n"
                f"🔢 Макс. использований: {max_uses}")
        await message.answer(text, reply_markup=admin_menu_keyboard(message.from_user.id), parse_mode="HTML")
        await state.clear()
    except ValueError:
        await message.answer("❌ Введите корректное число. Попробуйте снова:", parse_mode="HTML")


# ===================== АДМИН: УПРАВЛЕНИЕ АКЦИЯМИ =====================
@dp.callback_query(F.data == "admin_promotions")
async def cb_admin_promotions(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    promotions = await run_db(get_all_promotions)
    await run_db(add_admin_log, callback.from_user.id, "view_promotions", "Просмотрел список акций")
    text = "<b>🎉 Управление акциями</b>\n\n"
    if promotions:
        for promo in promotions:
            promo_id, name, description, discount, valid_until, is_active, created_at, service_id = promo
            status_icon = "🟢" if is_active else "🔴"
            valid_date = datetime.fromisoformat(valid_until).strftime("%d.%m.%Y") if valid_until else "∞"
            service_text = " (все услуги)" if service_id == 0 else f" (ID услуги: {service_id})"
            text += f"{status_icon} <b>{escape_html(name)}</b> - {discount}%{service_text} до {valid_date}\n"
            text += f"   ID: {promo_id}\n"
    else:
        text += "Нет созданных акций.\n"
    text += "\nВыберите действие:"
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="➕ Создать акцию", callback_data="promotion_create")
    keyboard.button(text="🔄 Активировать", callback_data="promotion_activate")
    keyboard.button(text="⏹️ Деактивировать", callback_data="promotion_deactivate")
    keyboard.button(text="🗑️ Удалить", callback_data="promotion_delete")
    keyboard.button(text="🔙 Назад", callback_data="admin_menu")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data == "promotion_create")
async def cb_promotion_create(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await callback.message.edit_text(
        "🎉 <b>Создание акции</b>\n\nВведите название акции (например: Новогодняя распродажа):",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminPromotionCreateState.waiting_for_name)
    await callback.answer()


@dp.message(AdminPromotionCreateState.waiting_for_name)
async def process_promotion_name(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Создание акции отменено.", reply_markup=admin_menu_keyboard(message.from_user.id),
                             parse_mode="HTML")
        return
    await state.update_data(name=message.text.strip())
    services = await run_db(get_all_services)
    text = "📝 <b>Выберите услугу для акции:</b>\n\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="🌍 На все услуги", callback_data="promotion_service_0")
    emoji_map = {
        "Курсовая работа": "📝",
        "Школьный проект": "🏫",
        "Отчёт по практике": "📊",
        "Доклад": "🎤",
        "Презентация": "📽️",
        "Защитное слово": "🛡️",
        "Реферат": "📚",
        "Редактирование работы": "✏️",
        "Тотальная защита (PREMIER)": "🛡️",
    }
    for s_id, name, desc, price, is_active in services:
        if is_active:
            emoji = emoji_map.get(name, "📋")
            keyboard.button(text=f"{emoji} {name}", callback_data=f"promotion_service_{s_id}")
    keyboard.button(text="🔙 Отмена", callback_data="admin_menu")
    keyboard.adjust(1)
    await message.answer(text, reply_markup=keyboard.as_markup(), parse_mode="HTML")
    await state.set_state(AdminPromotionCreateState.waiting_for_service)


@dp.callback_query(F.data.startswith("promotion_service_"))
async def cb_promotion_service(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    service_id = int(callback.data.split("_")[2])
    service_name = "все услуги"
    if service_id > 0:
        service = await run_db(get_service, service_id)
        if service:
            service_name = service[1]
    await state.update_data(service_id=service_id, service_name=service_name)
    await callback.message.edit_text(
        f"✅ Выбрано: <b>{service_name}</b>\n\n"
        f"📝 Введите <b>описание</b> акции (например: Скидка на курсовые работы в честь праздника!):",
        parse_mode="HTML"
    )
    await state.set_state(AdminPromotionCreateState.waiting_for_description)
    await callback.answer()


@dp.message(AdminPromotionCreateState.waiting_for_description)
async def process_promotion_description(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Создание акции отменено.", reply_markup=admin_menu_keyboard(message.from_user.id),
                             parse_mode="HTML")
        return
    await state.update_data(description=message.text.strip())
    await message.answer("🎯 Введите размер <b>скидки</b> в % (только число, например: 25):", parse_mode="HTML")
    await state.set_state(AdminPromotionCreateState.waiting_for_discount)


@dp.message(AdminPromotionCreateState.waiting_for_discount)
async def process_promotion_discount(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Создание акции отменено.", reply_markup=admin_menu_keyboard(message.from_user.id),
                             parse_mode="HTML")
        return
    try:
        discount = int(message.text.strip())
        if discount <= 0 or discount > 100:
            await message.answer("❌ Скидка должна быть от 1 до 100%. Попробуйте снова:", parse_mode="HTML")
            return
        await state.update_data(discount=discount)
        await message.answer(
            "📅 Введите дату окончания акции в формате <b>ДД.ММ.ГГГГ</b>\n"
            "Пример: 31.12.2026\n\n"
            "Или отправьте 'бесконечно', чтобы акция действовала без ограничений:",
            parse_mode="HTML"
        )
        await state.set_state(AdminPromotionCreateState.waiting_for_valid_until)
    except ValueError:
        await message.answer("❌ Введите корректное число. Попробуйте снова:", parse_mode="HTML")


@dp.message(AdminPromotionCreateState.waiting_for_valid_until)
async def process_promotion_valid_until(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Создание акции отменено.", reply_markup=admin_menu_keyboard(message.from_user.id),
                             parse_mode="HTML")
        return
    valid_until = None
    if message.text.strip().lower() == "бесконечно":
        valid_until = (datetime.now() + timedelta(days=365 * 100)).isoformat()
    else:
        try:
            valid_until_dt = datetime.strptime(message.text.strip(), "%d.%m.%Y")
            if valid_until_dt < datetime.now():
                await message.answer("❌ Дата должна быть в будущем. Попробуйте снова:", parse_mode="HTML")
                return
            valid_until = valid_until_dt.isoformat()
        except ValueError:
            await message.answer(
                "❌ Неверный формат. Используйте <b>ДД.ММ.ГГГГ</b> или напишите 'бесконечно'. Попробуйте снова:",
                parse_mode="HTML"
            )
            return
    data = await state.get_data()
    promo_id = await run_db(create_promotion, data['name'], data['description'], data['discount'], valid_until,
                            data.get('service_id', 0))
    await run_db(add_admin_log, message.from_user.id, "create_promotion",
                 f"Создал акцию {data['name']} ({data['discount']}%) для {data.get('service_name', 'всех услуг')}")
    valid_text = "бесконечно" if valid_until == (
            datetime.now() + timedelta(days=365 * 100)).isoformat() else datetime.fromisoformat(
        valid_until).strftime("%d.%m.%Y")
    service_text = "всем услугам" if data.get('service_id', 0) == 0 else data.get('service_name', 'всем услугам')
    await message.answer(
        f"✅ <b>Акция создана!</b>\n\n"
        f"📌 Название: {escape_html(data['name'])}\n"
        f"📝 Описание: {escape_html(data['description'])}\n"
        f"🎯 Скидка: <b>{data['discount']}%</b>\n"
        f"📅 Действует до: <b>{valid_text}</b>\n"
        f"🎯 Применяется к: <b>{service_text}</b>\n\n"
        f"Акция автоматически активна!",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )
    await state.clear()


@dp.callback_query(F.data == "promotion_activate")
async def cb_promotion_activate(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await callback.message.edit_text(
        "🔄 <b>Активация акции</b>\n\n"
        "Введите ID акции, которую хотите активировать.\n\n"
        "Чтобы узнать ID, посмотрите список акций выше.\n\n"
        "Пример: 1",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state("waiting_for_promo_activate")
    await callback.answer()


@dp.message(StateFilter("waiting_for_promo_activate"))
async def process_promotion_activate_id(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    try:
        promo_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введите корректный ID (только число).", parse_mode="HTML")
        return
    promo = await run_db(get_promotion_by_id, promo_id)
    if not promo:
        await message.answer(f"❌ Акция с ID {promo_id} не найдена.", parse_mode="HTML")
        await state.clear()
        return
    await run_db(activate_promotion, promo_id)
    await run_db(add_admin_log, message.from_user.id, "activate_promotion", f"Активировал акцию {promo[1]}")
    await message.answer(
        f"✅ Акция <b>{escape_html(promo[1])}</b> успешно активирована!",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )
    await state.clear()


@dp.callback_query(F.data == "promotion_deactivate")
async def cb_promotion_deactivate(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await callback.message.edit_text(
        "⏹️ <b>Деактивация акции</b>\n\n"
        "Введите ID акции, которую хотите деактивировать.\n\n"
        "Чтобы узнать ID, посмотрите список акций выше.\n\n"
        "Пример: 1",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state("waiting_for_promo_deactivate")
    await callback.answer()


@dp.message(StateFilter("waiting_for_promo_deactivate"))
async def process_promotion_deactivate_id(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    try:
        promo_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введите корректный ID (только число).", parse_mode="HTML")
        return
    promo = await run_db(get_promotion_by_id, promo_id)
    if not promo:
        await message.answer(f"❌ Акция с ID {promo_id} не найдена.", parse_mode="HTML")
        await state.clear()
        return
    await run_db(deactivate_promotion, promo_id)
    await run_db(add_admin_log, message.from_user.id, "deactivate_promotion", f"Деактивировал акцию {promo[1]}")
    await message.answer(
        f"✅ Акция <b>{escape_html(promo[1])}</b> успешно деактивирована!",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )
    await state.clear()


@dp.callback_query(F.data == "promotion_delete")
async def cb_promotion_delete(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await callback.message.edit_text(
        "🗑️ <b>Удаление акции</b>\n\n"
        "Введите ID акции, которую хотите удалить.\n\n"
        "Чтобы узнать ID, посмотрите список акций выше.\n\n"
        "Пример: 1\n\n"
        "⚠️ Это действие невозможно отменить!",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state("waiting_for_promo_delete")
    await callback.answer()


@dp.message(StateFilter("waiting_for_promo_delete"))
async def process_promotion_delete_id(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    try:
        promo_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введите корректный ID (только число).", parse_mode="HTML")
        return
    promo = await run_db(get_promotion_by_id, promo_id)
    if not promo:
        await message.answer(f"❌ Акция с ID {promo_id} не найдена.", parse_mode="HTML")
        await state.clear()
        return
    await run_db(delete_promotion, promo_id)
    await run_db(add_admin_log, message.from_user.id, "delete_promotion", f"Удалил акцию {promo[1]}")
    await message.answer(
        f"✅ Акция <b>{escape_html(promo[1])}</b> успешно удалена!",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )
    await state.clear()


# ===================== АДМИН: УПРАВЛЕНИЕ УСЛУГАМИ =====================
@dp.callback_query(F.data == "admin_services")
async def cb_admin_services(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await run_db(add_admin_log, callback.from_user.id, "view_services", "Просмотрел список услуг")
    services = await run_db(get_all_services)
    text = "🛠️ <b>Управление услугами</b>\n\nВыберите услугу для редактирования или добавьте новую:"
    await update_message(callback, text, get_services_keyboard(services), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("service_edit_") & ~F.data.startswith("service_edit_form_"))
async def cb_service_edit(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    try:
        service_id = int(callback.data.split("_")[2])
    except (ValueError, IndexError):
        await callback.answer("❌ Ошибка: неверный формат данных", show_alert=True)
        return
    service = await run_db(get_service, service_id)
    if not service:
        await callback.answer("❌ Услуга не найдена", show_alert=True)
        return
    _, name, description, price, is_active = service
    text = f"""
<b>🛠️ {escape_html(name)}</b>

📝 Описание: {escape_html(description)}
💰 Цена: <b>{price} ₽</b>
📊 Статус: {'✅ Активна' if is_active else '❌ Неактивна'}

Выберите действие:
"""
    await update_message(callback, text, service_edit_keyboard(service_id), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("service_edit_form_"))
async def cb_service_edit_form_start(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    try:
        service_id = int(callback.data.split("_")[3])
    except (ValueError, IndexError):
        await callback.answer("❌ Ошибка: неверный формат данных", show_alert=True)
        return
    service = await run_db(get_service, service_id)
    if not service:
        await callback.answer("❌ Услуга не найдена", show_alert=True)
        return
    await state.update_data(edit_service_id=service_id)
    await callback.message.edit_text(
        f"✏️ <b>Редактирование услуги</b>\n\n"
        f"Текущее название: {escape_html(service[1])}\n\n"
        f"Введите новое название (или отправьте /skip, чтобы оставить прежним):",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminServiceEditState.waiting_for_name)
    await callback.answer()


@dp.message(AdminServiceEditState.waiting_for_name)
async def process_service_edit_name(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return

    if message.text and message.text.strip() == "/skip":
        await state.update_data(new_name=None)
        await message.answer("✅ Название оставлено прежним. Введите новое описание (или /skip):", parse_mode="HTML")
        await state.set_state(AdminServiceEditState.waiting_for_description)
        return

    await state.update_data(new_name=message.text.strip())
    await message.answer("📝 Введите новое описание (или /skip, чтобы оставить прежним):", parse_mode="HTML")
    await state.set_state(AdminServiceEditState.waiting_for_description)


@dp.message(AdminServiceEditState.waiting_for_description)
async def process_service_edit_description(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return

    if message.text and message.text.strip() == "/skip":
        await state.update_data(new_description=None)
        await message.answer("✅ Описание оставлено прежним. Введите новую цену (или /skip):", parse_mode="HTML")
        await state.set_state(AdminServiceEditState.waiting_for_price)
        return

    if message.text.strip() != "/skip":
        await state.update_data(new_description=message.text.strip())
    await message.answer("💰 Введите новую цену числом (или /skip, чтобы оставить прежней):", parse_mode="HTML")
    await state.set_state(AdminServiceEditState.waiting_for_price)


@dp.message(AdminServiceEditState.waiting_for_price)
async def process_service_edit_price(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return

    data = await state.get_data()
    service_id = data.get("edit_service_id")
    service = await run_db(get_service, service_id)
    if not service:
        await message.answer("❌ Услуга не найдена.", reply_markup=admin_menu_keyboard(message.from_user.id),
                             parse_mode="HTML")
        await state.clear()
        return

    _, old_name, old_desc, old_price, is_active = service
    new_price = old_price

    if message.text and message.text.strip() == "/skip":
        new_price = old_price
    else:
        try:
            new_price = int(message.text.strip())
            if new_price <= 0:
                await message.answer("❌ Цена должна быть положительным числом. Попробуйте снова:", parse_mode="HTML")
                return
        except ValueError:
            await message.answer("❌ Введите корректное число или /skip. Попробуйте снова:", parse_mode="HTML")
            return

    new_name = data.get("new_name", old_name)
    new_description = data.get("new_description", old_desc)
    await run_db(update_service, service_id, new_name, new_description, new_price, is_active)
    await run_db(add_admin_log, message.from_user.id, "edit_service", f"Отредактировал услугу id={service_id}")
    await message.answer(
        f"✅ <b>Услуга обновлена!</b>\n\n📝 {escape_html(new_name)}\n💰 <b>{new_price} ₽</b>",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )
    await state.clear()


@dp.callback_query(F.data.startswith("service_toggle_"))
async def cb_service_toggle(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    service_id = int(callback.data.split("_")[2])
    service = await run_db(get_service, service_id)
    if not service:
        await callback.answer("❌ Услуга не найдена", show_alert=True)
        return
    _, name, description, price, is_active = service
    new_status = 0 if is_active else 1
    await run_db(update_service, service_id, name, description, price, new_status)
    await run_db(add_admin_log, callback.from_user.id, "toggle_service",
                 f"{'Активировал' if new_status else 'Деактивировал'} услугу {name}")
    await callback.answer(f"✅ Услуга {'активирована' if new_status else 'деактивирована'}!", show_alert=True)
    await cb_admin_services(callback)


@dp.callback_query(F.data.startswith("service_delete_"))
async def cb_service_delete(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    service_id = int(callback.data.split("_")[2])
    service = await run_db(get_service, service_id)
    if not service:
        await callback.answer("❌ Услуга не найдена", show_alert=True)
        return
    _, name, _, _, _ = service
    await run_db(delete_service, service_id)
    await run_db(add_admin_log, callback.from_user.id, "delete_service", f"Удалил услугу {name}")
    await callback.answer(f"✅ Услуга {escape_html(name)} удалена!", show_alert=True)
    await cb_admin_services(callback)


@dp.callback_query(F.data == "service_add")
async def cb_service_add(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await callback.message.edit_text(
        "➕ <b>Добавление новой услуги</b>\n\nВведите название услуги:",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminServiceAddState.waiting_for_name)
    await callback.answer()


@dp.message(AdminServiceAddState.waiting_for_name)
async def process_service_add_name(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    await state.update_data(name=message.text.strip())
    await message.answer("📝 Введите описание услуги:", parse_mode="HTML")
    await state.set_state(AdminServiceAddState.waiting_for_description)


@dp.message(AdminServiceAddState.waiting_for_description)
async def process_service_add_desc(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    await state.update_data(description=message.text.strip())
    await message.answer("💰 Введите цену услуги (только число):", parse_mode="HTML")
    await state.set_state(AdminServiceAddState.waiting_for_price)


@dp.message(AdminServiceAddState.waiting_for_price)
async def process_service_add_price(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    try:
        price = int(message.text.strip())
        if price <= 0:
            await message.answer("❌ Цена должна быть положительным числом. Попробуйте снова:", parse_mode="HTML")
            return
        data = await state.get_data()
        service_id = await run_db(add_service, data['name'], data['description'], price)
        await run_db(add_admin_log, message.from_user.id, "add_service", f"Добавил услугу {data['name']} ({price}₽)")
        await message.answer(
            f"✅ <b>Услуга {escape_html(data['name'])} успешно добавлена!</b>\n\n💰 Цена: <b>{price} ₽</b>",
            reply_markup=admin_menu_keyboard(message.from_user.id),
            parse_mode="HTML"
        )
        await state.clear()
    except ValueError:
        await message.answer("❌ Введите корректное число. Попробуйте снова:", parse_mode="HTML")


# ===================== АДМИН: УПРАВЛЕНИЕ ГОЛОСОВАНИЯМИ =====================
@dp.callback_query(F.data == "admin_polls")
async def cb_admin_polls(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    polls = await run_db(get_all_polls)
    await run_db(add_admin_log, callback.from_user.id, "view_polls", "Просмотрел список голосований")
    text = "🎯 <b>Управление голосованиями</b>\n\n"
    if not polls:
        text += "Нет созданных голосований."
    else:
        for poll in polls:
            poll_id, question, created_at, expires_at, is_active = poll
            status = "🟢 Активно" if is_active else "🔴 Завершено"
            created = datetime.fromisoformat(created_at).strftime("%d.%m %H:%M")
            text += f"• {status} - {escape_html(question[:30])}... [{created}]\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="➕ Создать голосование", callback_data="poll_create")
    keyboard.button(text="📊 Результаты", callback_data="poll_results")
    keyboard.button(text="🔙 Назад", callback_data="admin_menu")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data == "poll_create")
async def cb_poll_create(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    await callback.message.edit_text(
        "🎯 <b>Создание голосования</b>\n\nВведите вопрос для голосования:",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminPollCreateState.waiting_for_question)
    await callback.answer()


@dp.message(AdminPollCreateState.waiting_for_question)
async def process_poll_question(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    await state.update_data(question=message.text.strip())
    await message.answer(
        "📝 Введите варианты ответов через запятую\nПример: Да, Нет, Воздержался",
        parse_mode="HTML"
    )
    await state.set_state(AdminPollCreateState.waiting_for_options)


@dp.message(AdminPollCreateState.waiting_for_options)
async def process_poll_options(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    options = [opt.strip() for opt in message.text.split(",") if opt.strip()]
    if len(options) < 2:
        await message.answer("❌ Нужно минимум 2 варианта. Попробуйте снова:", parse_mode="HTML")
        return
    await state.update_data(options=options)
    await message.answer(
        "⏰ Введите срок действия в часах (например, 24):\nПо умолчанию - 24 часа",
        parse_mode="HTML"
    )
    await state.set_state(AdminPollCreateState.waiting_for_expiry)


@dp.message(AdminPollCreateState.waiting_for_expiry)
async def process_poll_expiry(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    try:
        hours = int(message.text.strip())
        if hours <= 0:
            hours = 24
    except ValueError:
        hours = 24
    data = await state.get_data()
    poll_id = await run_db(create_poll, data['question'], data['options'], message.from_user.id, hours)
    await run_db(add_admin_log, message.from_user.id, "create_poll", f"Создал голосование: {data['question']}")
    users = await run_db(get_all_users)
    text = (f"🎯 <b>НОВОЕ ГОЛОСОВАНИЕ!</b>\n\n"
            f"📋 {escape_html(data['question'])}\n\n"
            f"Варианты:\n" + "\n".join([f"• {escape_html(opt)}" for opt in data['options']]) +
            f"\n\n⏰ Голосование активно <b>{hours}</b> часов.")
    sent = 0
    for uid in users:
        try:
            await bot.send_message(uid[0], text, parse_mode="HTML")
            sent += 1
            await asyncio.sleep(0.05)
        except Exception as e:
            logging.debug(f"Не удалось отправить уведомление о голосовании {uid[0]}: {e}")
    await message.answer(
        f"✅ <b>Голосование создано!</b>\n"
        f"📋 {escape_html(data['question'])}\n"
        f"📝 Вариантов: {len(data['options'])}\n"
        f"⏰ {hours} часов\n"
        f"📨 Уведомлено пользователей: {sent}",
        reply_markup=admin_menu_keyboard(message.from_user.id),
        parse_mode="HTML"
    )
    await state.clear()


@dp.callback_query(F.data == "poll_results")
async def cb_poll_results(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    polls = await run_db(get_all_polls)
    if not polls:
        await update_message(callback, "📊 Нет созданных голосований.", admin_menu_keyboard(callback.from_user.id),
                             "HTML")
        await callback.answer()
        return
    text = "<b>📊 Результаты голосований:</b>\n\n"
    for poll in polls:
        poll_id, question, _, _, is_active = poll
        results = await run_db(get_poll_results, poll_id)
        status = "🟢 Активно" if is_active else "🔴 Завершено"
        text += f"{status} <b>{escape_html(question)}</b>\n"
        if results:
            total = sum(count for _, count in results)
            for option, count in results:
                percent = (count / total * 100) if total > 0 else 0
                text += f"  • {escape_html(option)}: {count} голосов ({percent:.1f}%)\n"
        else:
            text += "  • Нет голосов\n"
        text += "\n"
    await update_message(callback, text, admin_menu_keyboard(callback.from_user.id), "HTML")
    await callback.answer()


# ===================== АДМИН: УДАЛЕНИЕ СТАРЫХ ЗАКАЗОВ =====================
@dp.callback_query(F.data == "admin_delete_old")
async def cb_admin_delete_old(callback: CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    text = """
🗑️ <b>Удаление старых заказов</b>

Эта команда удаляет все оплаченные и отменённые заказы, которые старше указанного количества дней.

Введите количество дней (например, 30):
"""
    await callback.message.edit_text(text, reply_markup=back_to_admin_keyboard(), parse_mode="HTML")
    await state.set_state(AdminDeleteOldState.waiting_for_days)
    await callback.answer()


@dp.message(AdminDeleteOldState.waiting_for_days)
async def cb_admin_delete_old_process(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Только супер-админ имеет доступ.", parse_mode="HTML")
        await state.clear()
        return
    try:
        days = int(message.text.strip())
        if days <= 0:
            await message.answer("❌ Количество дней должно быть положительным числом. Попробуйте снова:",
                                 parse_mode="HTML")
            return
        deleted_count = await run_db(delete_old_orders, days)
        await run_db(add_admin_log, message.from_user.id, "delete_old_orders",
                     f"Удалил {deleted_count} заказов старше {days} дней")
        await message.answer(
            f"✅ Удалено <b>{deleted_count}</b> заказов, которые были старше <b>{days}</b> дней.\n\nУдалены только оплаченные и отменённые заказы.",
            reply_markup=admin_menu_keyboard(message.from_user.id), parse_mode="HTML")
        await state.clear()
    except ValueError:
        await message.answer("❌ Введите корректное число. Попробуйте снова:", parse_mode="HTML")


# ===================== АДМИН: ЭКСПОРТ ЗАКАЗОВ =====================
@dp.callback_query(F.data == "admin_export_orders")
async def cb_admin_export_orders(callback: CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Только супер-админ имеет доступ", show_alert=True)
        return
    if not EXCEL_AVAILABLE:
        await callback.answer("❌ Библиотека openpyxl не установлена.", show_alert=True)
        return
    await callback.answer("⏳ Формируем отчёт...")
    try:
        excel_file = await run_db(export_orders_to_excel)
        if not excel_file:
            await update_message(callback, "📊 Экспорт заказов\n\nНет заказов для экспорта.",
                                 admin_menu_keyboard(callback.from_user.id), "HTML")
            await callback.answer()
            return
        file_name = f"заказы_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
        await callback.message.answer_document(
            document=BufferedInputFile(excel_file.getvalue(), filename=file_name),
            caption=f"📊 <b>Отчёт по заказам</b>\n\n📅 Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n📦 Всего заказов: экспортировано\n\n🔐 Данные защищены и хранятся в децентрализованной системе.",
            parse_mode="HTML"
        )
        await update_message(callback, "✅ Отчёт успешно сформирован и отправлен!",
                             admin_menu_keyboard(callback.from_user.id), "HTML")
        await callback.answer()
    except Exception as e:
        logging.error(f"Ошибка экспорта: {e}")
        await callback.answer("❌ Ошибка при формировании отчёта", show_alert=True)


# ===================== АДМИН: ЛОГИ =====================
def _get_admin_logs(limit: int = 20):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT admin_id, action, details, timestamp FROM admin_logs ORDER BY timestamp DESC LIMIT ?", (limit,))
    rows = cur.fetchall()
    conn.close()
    return rows


@dp.callback_query(F.data == "admin_logs")
async def cb_admin_logs(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    await run_db(add_admin_log, callback.from_user.id, "view_logs", "Просмотрел логи")
    logs = await run_db(_get_admin_logs, 20)
    if not logs:
        text = "📋 Логов пока нет."
    else:
        text = "<b>📋 Последние действия администраторов:</b>\n\n"
        for admin_id, action, details, timestamp in logs:
            time_str = datetime.fromisoformat(timestamp).strftime("%d.%m %H:%M")
            text += f"• {time_str} - {escape_html(action)} {escape_html(details)}\n"
    await update_message(callback, text, admin_menu_keyboard(callback.from_user.id), "HTML")
    await callback.answer()


# ===================== АДМИН: НАЗНАЧЕНИЕ ЦЕНЫ =====================
@dp.callback_query(F.data.startswith("set_price_"))
async def cb_set_price_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    order_id = int(callback.data.split("_")[2])
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    await state.update_data(order_id=order_id)
    await callback.message.edit_text(
        f"💰 <b>Назначение цены для заказа {order[9] or f'#{order_id}'}</b>\n\nТекущая цена: {order[3]} руб.\n\nВведите новую цену (только число):",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminSetPriceState.waiting_for_price)
    await callback.answer()


@dp.message(AdminSetPriceState.waiting_for_price)
async def cb_set_price_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    try:
        new_price = int(message.text.strip())
        if new_price <= 0:
            await message.answer("❌ Цена должна быть положительным числом. Попробуйте снова:", parse_mode="HTML")
            return
        data = await state.get_data()
        order_id = data.get("order_id")
        order = await run_db(get_order, order_id)
        order_code = order[9] if order else f"#{order_id}"
        await run_db(update_order_price, order_id, new_price, "")
        await run_db(add_admin_log, message.from_user.id, "set_price",
                     f"Назначил цену {new_price} руб. для заказа {order_code}")
        await message.answer(
            f"✅ Цена для заказа <b>{escape_html(order_code)}</b> успешно обновлена на <b>{new_price} руб.</b>\n\n"
            f"Теперь вы можете подтвердить оплату!",
            parse_mode="HTML")
        await state.clear()
        await send_order_detail_message(message, order_id)
    except ValueError:
        await message.answer("❌ Введите корректное число. Попробуйте снова:", parse_mode="HTML")


async def send_order_detail_message(message: Message, order_id: int):
    order = await run_db(get_order, order_id)
    if not order:
        await message.answer("❌ Заказ не найден", parse_mode="HTML")
        return
    user = await run_db(get_user, order[1])
    user_display = f"@{user[1]}" if user and user[1] else f"ID:{order[1]}"

    order_id, user_id, service, price, status, created_at, paid_at, admin_price, admin_note, order_code, rating, review, file_id, is_urgent = order
    status_text = "✅ Оплачен" if status == "paid" else "⏳ Ожидает оплаты" if status == "pending" else "🔧 В работе" if status == "in_progress" else "❌ Отменён"
    final_price = admin_price if admin_price > 0 else price
    display_code = order_code or f"#{order_id}"
    urgent = "🔥 Срочный заказ!\n" if is_urgent else ""
    created = datetime.fromisoformat(created_at).strftime("%d.%m.%Y %H:%M")
    paid = datetime.fromisoformat(paid_at).strftime("%d.%m.%Y %H:%M") if paid_at else "Не оплачен"

    text = f"""
<b>📦 Информация о заказе {escape_html(display_code)}</b>

{urgent}👤 Пользователь: {escape_html(user_display)} (ID: {user_id})
📝 Услуга: {escape_html(service)}
💰 Изначальная цена: {price} руб.
💰 Назначенная цена: <b>{final_price} руб.</b>
📊 Статус: {status_text}
📅 Создан: {created}
✅ Оплачен: {paid}
"""
    if rating > 0:
        text += f"⭐ Оценка: {rating}/5\n"
        if review:
            text += f"📝 Отзыв: {escape_html(review)}\n"
    if file_id:
        text += f"📎 Прикреплён файл: ✅\n"
    if admin_note:
        text += f"📌 Заметка: {escape_html(admin_note)}\n"

    await message.answer(text, reply_markup=order_detail_keyboard(order_id, status, is_urgent),
                         parse_mode="HTML")


# ===================== АДМИН: ПОЛЬЗОВАТЕЛИ =====================
@dp.callback_query(F.data == "admin_users")
async def cb_admin_users(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    users = await run_db(get_all_users)
    await run_db(add_admin_log, callback.from_user.id, "view_users", f"Просмотрел список пользователей ({len(users)})")
    if not users:
        await update_message(callback, "👥 Пользователей пока нет.", admin_menu_keyboard(callback.from_user.id), "HTML")
        await callback.answer()
        return
    await update_message(callback, "👥 <b>Список пользователей:</b>", users_keyboard(users, 0), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("users_page_"))
async def cb_users_page(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    page = int(callback.data.split("_")[2])
    users = await run_db(get_all_users)
    await callback.message.edit_reply_markup(reply_markup=users_keyboard(users, page))
    await callback.answer()


@dp.callback_query(F.data.startswith("user_orders_"))
async def cb_user_orders_detail(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    user_id = int(callback.data.split("_")[2])
    orders = await run_db(get_user_orders, user_id)
    if not orders:
        text = "📦 У пользователя нет заказов."
    else:
        text = f"📦 <b>Заказы пользователя (ID: {user_id}):</b>\n\n"
        for order in orders:
            order_id, service, price, status, created_at, admin_price, order_code, rating, _, is_urgent = order
            status_text = "✅ Оплачен" if status == "paid" else "⏳ Ожидает" if status == "pending" else "🔧 В работе" if status == "in_progress" else "❌ Отменён"
            final_price = admin_price if admin_price > 0 else price
            created = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
            display_code = order_code or f"#{order_id}"
            urgent = "🔥 " if is_urgent else ""
            rating_str = f"⭐{rating}" if rating > 0 else ""
            text += f"• {urgent}{escape_html(display_code)}: {escape_html(service)} - {final_price} руб. ({status_text}) [{created}] {rating_str}\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="🔙 Назад", callback_data=f"user_{user_id}")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("user_") & ~F.data.startswith("user_orders_"))
async def cb_user_detail(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    try:
        user_id = int(callback.data.split("_")[1])
    except (ValueError, IndexError):
        await callback.answer("❌ Ошибка", show_alert=True)
        return
    user = await run_db(get_user, user_id)
    if not user:
        await callback.answer("❌ Пользователь не найден", show_alert=True)
        return
    _, username, first_name, last_name, reg_date, birthday, used_promocodes = user
    logs = await run_db(get_user_logs, user_id)
    user_stats = await run_db(get_user_stats, user_id)
    used_list = used_promocodes.split(",") if used_promocodes else []
    text = f"""
<b>👤 Информация о пользователе</b>

🆔 ID: {user_id}
👤 Имя: {escape_html(first_name)} {escape_html(last_name or '')}
📌 Username: @{escape_html(username or 'Не указан')}
📅 Регистрация: {datetime.fromisoformat(reg_date).strftime('%d.%m.%Y')}
🎂 День рождения: {birthday or 'Не указан'}

📦 Заказов: {user_stats['total_orders']}
✅ Оплачено: {user_stats['paid_orders']}
🔧 В работе: {user_stats['in_progress']}
💰 Потрачено: {user_stats['total_spent']} руб.

🏷️ Использованные промокоды: {', '.join(used_list) if used_list else 'Нет'}

<b>📋 Последние действия:</b>
"""
    for action, details, timestamp in logs[:5]:
        time_str = datetime.fromisoformat(timestamp).strftime("%d.%m %H:%M")
        text += f"• {time_str} - {escape_html(action)} {escape_html(details)}\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="📦 Заказы пользователя", callback_data=f"user_orders_{user_id}")
    keyboard.button(text="🔙 Назад", callback_data="admin_users")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


# ===================== АДМИН: ЗАКАЗЫ =====================
@dp.callback_query(F.data == "admin_orders")
async def cb_admin_orders(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    orders = await run_db(get_all_orders)
    await run_db(add_admin_log, callback.from_user.id, "view_orders", f"Просмотрел список заказов ({len(orders)})")
    if not orders:
        await update_message(callback, "📦 Заказов пока нет.", admin_menu_keyboard(callback.from_user.id), "HTML")
        await callback.answer()
        return
    await update_message(callback, "📦 <b>Список заказов:</b>", orders_keyboard(orders, 0), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("orders_page_"))
async def cb_orders_page(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    page = int(callback.data.split("_")[2])
    orders = await run_db(get_all_orders)
    await callback.message.edit_reply_markup(reply_markup=orders_keyboard(orders, page))
    await callback.answer()


@dp.callback_query(F.data.startswith("order_"))
async def cb_order_detail(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    order_id = int(callback.data.split("_")[1])
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    user = await run_db(get_user, order[1])
    user_display = f"@{user[1]}" if user and user[1] else f"ID:{order[1]}"

    order_id, user_id, service, price, status, created_at, paid_at, admin_price, admin_note, order_code, rating, review, file_id, is_urgent = order
    status_text = "✅ Оплачен" if status == "paid" else "⏳ Ожидает оплаты" if status == "pending" else "🔧 В работе" if status == "in_progress" else "❌ Отменён"
    final_price = admin_price if admin_price > 0 else price
    display_code = order_code or f"#{order_id}"
    urgent = "🔥 Срочный заказ!\n" if is_urgent else ""
    created = datetime.fromisoformat(created_at).strftime("%d.%m.%Y %H:%M")
    paid = datetime.fromisoformat(paid_at).strftime("%d.%m.%Y %H:%M") if paid_at else "Не оплачен"

    text = f"""
<b>📦 Информация о заказе {escape_html(display_code)}</b>

{urgent}👤 Пользователь: {escape_html(user_display)} (ID: {user_id})
📝 Услуга: {escape_html(service)}
💰 Изначальная цена: {price} руб.
💰 Назначенная цена: <b>{final_price} руб.</b>
📊 Статус: {status_text}
📅 Создан: {created}
✅ Оплачен: {paid}
"""
    if rating > 0:
        text += f"⭐ Оценка: {rating}/5\n"
        if review:
            text += f"📝 Отзыв: {escape_html(review)}\n"
    if file_id:
        text += f"📎 Прикреплён файл: ✅\n"
    if admin_note:
        text += f"📌 Заметка: {escape_html(admin_note)}\n"

    await update_message(callback, text, order_detail_keyboard(order_id, status, is_urgent), "HTML")
    await callback.answer()


# ===================== АДМИН: ПРИНЯТЬ СРОЧНЫЙ ЗАКАЗ =====================
@dp.callback_query(F.data.startswith("accept_urgent_"))
async def cb_accept_urgent(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    order_id = int(callback.data.split("_")[2])
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    if order[4] != "pending":
        await callback.answer("❌ Заказ уже обработан", show_alert=True)
        return
    await run_db(update_order_status, order_id, "in_progress")
    await run_db(add_admin_log, callback.from_user.id, "accept_urgent", f"Принял срочный заказ {order[9]}")
    try:
        await bot.send_message(order[1],
                               f"🔥 <b>Срочный заказ принят!</b>\n\nЗаказ {order[9]}: {order[2]}\nСтатус: <b>В работе</b>\n\nВаш срочный заказ взят в работу!",
                               parse_mode="HTML")
    except Exception as e:
        logging.warning(f"Не удалось уведомить пользователя {order[1]}: {e}")
    await callback.answer("✅ Срочный заказ принят в работу!", show_alert=True)
    await cb_order_detail(callback)


# ===================== АДМИН: ОТЗЫВЫ =====================
@dp.callback_query(F.data == "admin_reviews")
async def cb_admin_reviews(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    reviews = await run_db(get_all_reviews)
    await run_db(add_admin_log, callback.from_user.id, "view_reviews", "Просмотрел список отзывов")
    if not reviews:
        await update_message(callback, "⭐ Отзывов пока нет.", admin_menu_keyboard(callback.from_user.id), "HTML")
        await callback.answer()
        return
    await update_message(callback, "⭐ <b>Управление отзывами</b>\n\nВыберите отзыв для управления:",
                         reviews_keyboard(reviews, 0), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("reviews_page_"))
async def cb_reviews_page(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    page = int(callback.data.split("_")[2])
    reviews = await run_db(get_all_reviews)
    await callback.message.edit_reply_markup(reply_markup=reviews_keyboard(reviews, page))
    await callback.answer()


def _get_review_by_code(order_code: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT o.order_code, o.service, o.rating, o.review, u.username, u.first_name, o.created_at, o.order_id
        FROM orders o
        LEFT JOIN users u ON o.user_id = u.user_id
        WHERE o.order_code = ? AND o.rating > 0
    """, (order_code,))
    row = cur.fetchone()
    conn.close()
    return row


def _delete_review_by_code(order_code: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT order_id, service, rating, review FROM orders WHERE order_code = ?", (order_code,))
    row = cur.fetchone()
    conn.close()
    return row


def _reset_review_by_code(order_code: str):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE orders SET rating = 0, review = '' WHERE order_code = ?", (order_code,))
    conn.commit()
    conn.close()


@dp.callback_query(F.data.startswith("review_detail_"))
async def cb_review_detail(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    order_code = callback.data.split("_")[2]
    review = await run_db(_get_review_by_code, order_code)
    if not review:
        await callback.answer("❌ Отзыв не найден", show_alert=True)
        return
    order_code, service, rating, review_text, username, first_name, created_at, order_id = review
    name = username or first_name or "Аноним"
    created = datetime.fromisoformat(created_at).strftime("%d.%m.%Y %H:%M")
    stars = "⭐" * rating + "☆" * (5 - rating)
    text = f"""
<b>⭐ Детали отзыва</b>

📌 Заказ: {escape_html(order_code)}
📝 Услуга: {escape_html(service)}
👤 Автор: {escape_html(name)}
{stars} {rating}/5
📝 Отзыв: {escape_html(review_text or 'Без текста')}
📅 Дата: {created}
"""
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="❌ Удалить отзыв", callback_data=f"delete_review_{order_code}")
    keyboard.button(text="🔙 Назад", callback_data="admin_reviews")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("delete_review_"))
async def cb_delete_review_confirm(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    order_code = callback.data.split("_")[2]
    review = await run_db(_delete_review_by_code, order_code)
    if not review:
        await callback.answer("❌ Отзыв не найден", show_alert=True)
        return
    order_id, service, rating, review_text = review
    text = f"""
<b>⚠️ Вы уверены, что хотите удалить этот отзыв?</b>

📌 Заказ: {escape_html(order_code)}
📝 Услуга: {escape_html(service)}
⭐ Оценка: {rating}/5
📝 Отзыв: {escape_html(review_text or 'Без текста')}

Это действие невозможно отменить!
"""
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="✅ Да, удалить", callback_data=f"confirm_delete_review_{order_code}")
    keyboard.button(text="❌ Отмена", callback_data=f"review_detail_{order_code}")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("confirm_delete_review_"))
async def cb_confirm_delete_review(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    order_code = callback.data.split("_")[3]
    await run_db(_reset_review_by_code, order_code)
    await run_db(add_admin_log, callback.from_user.id, "delete_review", f"Удалил отзыв к заказу {order_code}")
    await callback.message.edit_text(f"✅ Отзыв к заказу <b>{escape_html(order_code)}</b> успешно удалён!",
                                     reply_markup=admin_menu_keyboard(callback.from_user.id), parse_mode="HTML")
    await callback.answer()


# ===================== АДМИН: ПРИКРЕПИТЬ ФАЙЛ =====================
@dp.callback_query(F.data.startswith("attach_file_"))
async def cb_attach_file_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    order_id = int(callback.data.split("_")[2])
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    await state.update_data(order_id=order_id)
    await callback.message.edit_text(
        f"📎 <b>Прикрепить файл к заказу {order[9] or f'#{order_id}'}</b>\n\nОтправьте файл (PDF, DOC, DOCX, TXT, ZIP, JPG, PNG):",
        reply_markup=back_to_admin_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AttachFileState.waiting_for_file)
    await callback.answer()


@dp.message(AttachFileState.waiting_for_file)
async def cb_attach_file_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа.", parse_mode="HTML")
        await state.clear()
        return
    data = await state.get_data()
    order_id = data.get("order_id")
    if not order_id:
        await message.answer("❌ Ошибка: заказ не найден.", parse_mode="HTML")
        await state.clear()
        return
    order = await run_db(get_order, order_id)
    order_code = order[9] if order else f"#{order_id}"
    if message.document:
        file_id = message.document.file_id
        file_name = message.document.file_name or "файл"
    elif message.photo:
        file_id = message.photo[-1].file_id
        file_name = "фото.jpg"
    else:
        await message.answer("❌ Пожалуйста, отправьте файл (PDF, DOC, DOCX, TXT, ZIP, JPG, PNG).", parse_mode="HTML")
        return
    await run_db(update_order_file, order_id, file_id)
    await run_db(add_admin_log, message.from_user.id, "attach_file", f"Прикрепил файл к заказу {order_code}")
    await message.answer(
        f"✅ Файл <b>{escape_html(file_name)}</b> успешно прикреплён к заказу <b>{escape_html(order_code)}</b>!",
        reply_markup=admin_menu_keyboard(message.from_user.id), parse_mode="HTML")
    await state.clear()
    if order:
        try:
            await bot.send_message(order[1],
                                   f"📎 <b>К вашему заказу прикреплён файл!</b>\n\nЗаказ {order_code}: {order[2]}\nФайл: {file_name}\n\nВы можете скачать его в разделе 'Мои заказы'.",
                                   parse_mode="HTML")
        except Exception as e:
            logging.warning(f"Не удалось уведомить пользователя {order[1]}: {e}")


# ===================== ОТЗЫВЫ (ПОЛЬЗОВАТЕЛЬ) =====================
@dp.callback_query(F.data.startswith("review_order_"))
async def cb_review_start(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    order_id = int(callback.data.split("_")[2])
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    if order[1] != user_id:
        await callback.answer("⛔ Это не ваш заказ", show_alert=True)
        return
    if order[4] != "paid":
        await callback.answer("❌ Оставить отзыв можно только для выполненного заказа", show_alert=True)
        return
    if order[10] > 0:
        await callback.answer("❌ Вы уже оставили отзыв на этот заказ", show_alert=True)
        return
    await state.update_data(order_id=order_id)
    keyboard = InlineKeyboardBuilder()
    for i in range(1, 6):
        keyboard.button(text=f"⭐ {i}", callback_data=f"rating_{i}")
    keyboard.button(text="❌ Отмена", callback_data="my_orders")
    keyboard.adjust(5, 1)
    await callback.message.edit_text(
        f"⭐ <b>Оцените работу над заказом {order[9] or f'#{order_id}'}</b>\n\nВыберите оценку от 1 до 5:",
        reply_markup=keyboard.as_markup(), parse_mode="HTML")
    await state.set_state(ReviewState.waiting_for_rating)
    await callback.answer()


@dp.callback_query(F.data.startswith("rating_"))
async def cb_review_rating(callback: CallbackQuery, state: FSMContext):
    rating = int(callback.data.split("_")[1])
    await state.update_data(rating=rating)
    await callback.message.edit_text(
        f"⭐ <b>Оставьте отзыв</b>\n\nВы выбрали оценку: <b>{rating}/5</b>\n\nНапишите текст отзыва (или отправьте /skip, чтобы пропустить):",
        parse_mode="HTML")
    await state.set_state(ReviewState.waiting_for_review)
    await callback.answer()


@dp.message(ReviewState.waiting_for_review)
async def cb_review_process(message: Message, state: FSMContext):
    if message.text and message.text == "/skip":
        review_text = ""
    else:
        review_text = message.text
    data = await state.get_data()
    order_id = data.get("order_id")
    rating = data.get("rating")
    if not order_id:
        await message.answer("❌ Ошибка: заказ не найден.", parse_mode="HTML")
        await state.clear()
        return
    await run_db(update_order_review, order_id, rating, review_text)
    order = await run_db(get_order, order_id)
    order_code = order[9] if order else f"#{order_id}"
    await run_db(add_user_log, message.from_user.id, "review", f"Оставил отзыв на заказ {order_code}: {rating}/5")
    await message.answer(
        f"✅ <b>Спасибо за ваш отзыв!</b>\n\n⭐ Оценка: <b>{rating}/5</b>\n📝 Отзыв: {escape_html(review_text or 'Без текста')}\n\nМы ценим ваше мнение!",
        reply_markup=back_to_main_keyboard(), parse_mode="HTML")
    await state.clear()


@dp.message(StateFilter(ReviewState.waiting_for_review), F.text == "/cancel")
async def cb_review_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Отзыв отменён.", reply_markup=main_menu_keyboard(), parse_mode="HTML")


# ===================== ПОЛЬЗОВАТЕЛЬ: ОТМЕНА ЗАКАЗА =====================
@dp.callback_query(F.data.startswith("cancel_order_"))
async def cb_cancel_order(callback: CallbackQuery):
    order_id = int(callback.data.split("_")[2])
    user_id = callback.from_user.id
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    if order[1] != user_id:
        await callback.answer("⛔ Это не ваш заказ", show_alert=True)
        return
    if order[4] == "paid":
        await callback.answer("❌ Оплаченный заказ нельзя отменить", show_alert=True)
        return
    if order[4] == "cancelled":
        await callback.answer("❌ Заказ уже отменён", show_alert=True)
        return
    if order[4] == "in_progress":
        await callback.answer("❌ Заказ уже в работе, отмена невозможна", show_alert=True)
        return
    order_code = order[9] or f"#{order_id}"
    await run_db(update_order_status, order_id, "cancelled")
    await run_db(add_user_log, user_id, "cancel_order", f"Отменил заказ {order_code}")
    await callback.message.edit_text(
        f"✅ Заказ <b>{escape_html(order_code)}</b> успешно отменён!\n\nЕсли вы передумали, вы можете создать новый заказ.",
        reply_markup=back_to_main_keyboard(), parse_mode="HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("refresh_order_"))
async def cb_refresh_order(callback: CallbackQuery):
    await cb_user_order_detail(callback)


# ===================== ПОЛЬЗОВАТЕЛЬ: ДЕТАЛИ ЗАКАЗА =====================
@dp.callback_query(F.data.startswith("my_order_"))
async def cb_user_order_detail(callback: CallbackQuery):
    order_id = int(callback.data.split("_")[2])
    user_id = callback.from_user.id
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    if order[1] != user_id:
        await callback.answer("⛔ Это не ваш заказ", show_alert=True)
        return
    order_id, user_id, service, price, status, created_at, paid_at, admin_price, admin_note, order_code, rating, review, file_id, is_urgent = order
    status_text = "✅ Оплачен" if status == "paid" else "⏳ Ожидает оплаты" if status == "pending" else "🔧 В работе" if status == "in_progress" else "❌ Отменён"
    final_price = admin_price if admin_price > 0 else price
    display_code = order_code or f"#{order_id}"
    urgent = "🔥 Срочный заказ!\n" if is_urgent else ""
    created = datetime.fromisoformat(created_at).strftime("%d.%m.%Y %H:%M")
    paid = datetime.fromisoformat(paid_at).strftime("%d.%m.%Y %H:%M") if paid_at else "Не оплачен"
    text = f"""
<b>📦 Заказ {escape_html(display_code)}</b>

{urgent}📝 Услуга: {escape_html(service)}
💰 Цена: <b>{final_price} руб.</b>
📊 Статус: {status_text}
📅 Создан: {created}
✅ Оплачен: {paid}
"""
    if rating > 0:
        text += f"⭐ Оценка: {rating}/5\n"
        if review:
            text += f"📝 Отзыв: {escape_html(review)}\n"
    if file_id:
        text += f"\n📎 <b>Файл прикреплён!</b>\n"
        keyboard = InlineKeyboardBuilder()
        keyboard.button(text="📥 Скачать файл", callback_data=f"download_file_{order_id}")
        keyboard.button(text="🔙 Назад", callback_data="my_orders")
        keyboard.adjust(1)
        await update_message(callback, text, keyboard.as_markup(), "HTML")
        await callback.answer()
        return
    if admin_note:
        text += f"\n📌 Заметка: {escape_html(admin_note)}\n"

    await update_message(callback, text, order_user_keyboard(order_id, status), "HTML")
    await callback.answer()


# ===================== СКАЧИВАНИЕ ФАЙЛА =====================
@dp.callback_query(F.data.startswith("download_file_"))
async def cb_download_file(callback: CallbackQuery):
    user_id = callback.from_user.id
    order_id = int(callback.data.split("_")[2])
    order = await run_db(get_order, order_id)
    if not order:
        await callback.answer("❌ Заказ не найден", show_alert=True)
        return
    if order[1] != user_id and not is_admin(user_id):
        await callback.answer("⛔ Нет доступа", show_alert=True)
        return
    file_id = order[12]
    if not file_id:
        await callback.answer("❌ Файл не найден", show_alert=True)
        return
    try:
        await bot.send_document(user_id, file_id,
                                caption=f"📎 Файл к заказу {order[9] or f'#{order_id}'}\nУслуга: {order[2]}")
        await callback.answer("✅ Файл отправлен!")
    except Exception as e:
        logging.error(f"Ошибка отправки файла: {e}")
        await callback.answer("❌ Ошибка при отправке файла", show_alert=True)


# ===================== ПОКУПКА (ВЫБОР УСЛУГИ) =====================
@dp.callback_query(F.data == "buy")
async def cb_buy(callback: CallbackQuery):
    services = await run_db(get_all_services)
    if not services:
        await update_message(callback,
                             "📚 <b>Услуги</b>\n\nК сожалению, в данный момент услуги не доступны. Пожалуйста, свяжитесь с поддержкой.",
                             back_to_main_keyboard(), "HTML")
        await callback.answer()
        return
    await update_message(callback, "📚 <b>Выберите услугу:</b>",
                         services_keyboard_from_db(services, callback.from_user.id), "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("buyservice_"))
async def cb_service_from_db(callback: CallbackQuery, state: FSMContext):
    service_id = int(callback.data.split("_")[1])
    service = await run_db(get_service, service_id)
    if not service:
        await callback.answer("❌ Услуга не найдена", show_alert=True)
        return
    _, name, description, price, is_active = service
    if not is_active:
        await callback.answer("❌ Эта услуга временно недоступна", show_alert=True)
        return

    promotions = await run_db(get_active_promotions)
    promo_discount = 0
    promo_name = ""
    for promo in promotions:
        if promo[3] > promo_discount:
            promo_discount = promo[3]
            promo_name = promo[1]

    discount, discount_code = await run_db(get_pending_discount, callback.from_user.id)
    final_discount = max(promo_discount, discount)
    final_discount_name = discount_code if discount >= promo_discount else promo_name

    await state.update_data(service_id=service_id, service_name=name, service_price=price, discount=final_discount,
                            discount_name=final_discount_name)

    text = f"""
<b>📋 Вы выбрали: {escape_html(name)}</b>

📝 {escape_html(description)}

💰 <b>Цена:</b> {format_price_with_discount(price, final_discount)}
"""
    if final_discount > 0:
        text += f"\n🎉 <b>Применяется скидка: {final_discount}%</b> (акция: {escape_html(final_discount_name)})"
    text += """

📌 <i>Важно! Окончательная цена и сроки зависят от:</i>
• Тема работы
• Сложность и объём
• Текущая загрузка команды

✅ <b>Подтвердите создание заказа:</b>
"""
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="✅ Обычный заказ", callback_data=f"confirm_order_{service_id}")
    keyboard.button(text="🔥 Срочный заказ", callback_data=f"confirm_order_urgent_{service_id}")
    keyboard.button(text="❌ Отмена", callback_data="buy")
    keyboard.adjust(1)
    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()


# ===================== ОБЫЧНЫЙ ЗАКАЗ =====================
@dp.callback_query(F.data.startswith("confirm_order_"))
async def cb_confirm_order_from_db(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    parts = callback.data.split("_")

    is_urgent = 0
    if len(parts) >= 4 and parts[2] == "urgent":
        is_urgent = 1
        service_id = int(parts[3])
    else:
        service_id = int(parts[2])

    data = await state.get_data()
    service_name = data.get("service_name")
    service_price = data.get("service_price")
    discount = data.get("discount", 0)
    discount_name = data.get("discount_name", "")

    if not service_name:
        await callback.answer("❌ Ошибка: выберите услугу заново", show_alert=True)
        await state.clear()
        return

    final_price = max(1, round(service_price * (1 - discount / 100))) if discount > 0 else service_price
    if discount > 0:
        await run_db(clear_pending_discount, user_id)

    order_id, order_code = await run_db(add_order, user_id, service_name, final_price, is_urgent, discount)
    await run_db(update_user_action, user_id, f"order_{service_name}")
    await run_db(add_user_log, user_id, "create_order",
                 f"Заказ {order_code}: {service_name} ({final_price}₽) {'СРОЧНЫЙ' if is_urgent else ''} {'СКИДКА ' + str(discount) + '%' if discount else ''}")

    await state.clear()

    user = await run_db(get_user, user_id)
    username = user[1] if user else None
    user_name = user[2] if user else "Пользователь"

    urgent_text = "🔥 <b>СРОЧНЫЙ ЗАКАЗ!</b>\n" if is_urgent else ""

    if discount > 0:
        price_line = f"💰 Стоимость со скидкой: <b>{final_price} руб.</b> (базовая {service_price} руб., скидка <b>{discount}%</b> по акции {escape_html(discount_name)})\n"
    else:
        price_line = f"💰 Стоимость: <b>{final_price} руб.</b>\n"

    text = f"""
<b>✅ Заказ успешно создан!</b>

{urgent_text}📋 Услуга: <b>{escape_html(service_name)}</b>
{price_line}🏷️ Код заказа: <b>{escape_html(order_code)}</b>

📌 <i>Важно! Окончательная цена и сроки зависят от:</i>
• Тема работы
• Сложность и объём
• Текущая загрузка команды

📞 Для уточнения стоимости и оформления заказа свяжитесь с диспетчером:
{DISPATCHER_USERNAME}
👤 Или с CEO: {CEO_USERNAME}

💬 После согласования всех деталей сообщите диспетчеру код заказа: <b>{escape_html(order_code)}</b>

🔐 <i>Ваш заказ защищён и хранится в децентрализованной системе.</i>
"""

    keyboard = InlineKeyboardBuilder()
    dispatcher_username = DISPATCHER_USERNAME.replace("@", "")
    ceo_username = CEO_USERNAME.replace("@", "")
    keyboard.button(text="📞 Связаться с диспетчером", url=f"https://t.me/{dispatcher_username}")
    keyboard.button(text="👤 Связаться с CEO", url=f"https://t.me/{ceo_username}")
    keyboard.button(text="📋 Мои заказы", callback_data="my_orders")
    keyboard.button(text="🔙 На главную", callback_data="main_menu")
    keyboard.adjust(1)

    await update_message(callback, text, keyboard.as_markup(), "HTML")
    await callback.answer()

    for admin_id in ADMINS:
        try:
            urgent_marker = "🔥 СРОЧНЫЙ " if is_urgent else ""
            msg = "🆕 " + urgent_marker + "НОВЫЙ ЗАКАЗ!"
            msg += "\n📋 Услуга: " + service_name
            msg += "\n🏷️ Код: " + order_code
            msg += "\n👤 Пользователь: @" + (username or "без username") + " (" + user_name + ")"
            msg += "\n💰 Цена: " + str(final_price) + " ₽"
            if discount > 0:
                msg += f"\n🎉 Скидка {discount}% по акции {discount_name} (база {service_price}₽)"
            if is_urgent:
                msg += "\n🔥 Срочный заказ требует немедленного внимания!"
            msg += "\n📅 " + datetime.now().strftime('%d.%m.%Y %H:%M')
            await bot.send_message(admin_id, msg)
        except Exception as e:
            logging.error(f"Ошибка отправки уведомления админу {admin_id}: {e}")

    if is_urgent:
        asyncio.create_task(
            urgent_notification_loop(order_id, order_code, service_name, username, user_name, final_price))


async def urgent_notification_loop(order_id: int, order_code: str, service_name: str, username: str, user_name: str,
                                   price: int):
    attempts = 0
    max_attempts = 30
    while attempts < max_attempts:
        await asyncio.sleep(120)
        order = await run_db(get_order, order_id)
        if not order:
            break
        if order[4] != "pending":
            break
        attempts += 1
        for admin_id in ADMINS:
            try:
                await bot.send_message(admin_id,
                                       f"🔥 <b>СРОЧНЫЙ ЗАКАЗ ОЖИДАЕТ!</b>\n\n📋 Услуга: {service_name}\n🏷️ Код: {order_code}\n👤 Пользователь: @{username or 'без username'} ({user_name})\n💰 Цена: <b>{price} ₽</b>\n⏰ Заказ ожидает уже <b>{attempts * 2}</b> минут\n\n⚠️ Нажмите на заказ и выберите 'Принять срочный заказ'!",
                                       parse_mode="HTML")
            except Exception as e:
                logging.warning(f"Не удалось отправить напоминание админу {admin_id}: {e}")


# ===================== ПОДДЕРЖКА =====================
@dp.callback_query(F.data == "support")
async def cb_support(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "support")
    await run_db(add_user_log, user_id, "support", "Открыл поддержку")
    text = f"""
<b>📞 Техническая поддержка</b>

Напишите ваше сообщение, и я перешлю его автору. Автор ответит вам в этом же чате.

Также вы можете связаться с диспетчером:
{DISPATCHER_USERNAME}
👤 Или с CEO: {CEO_USERNAME}

Для выхода из режима поддержки отправьте /cancel.
"""
    await update_message(callback, text, back_to_main_keyboard(), "HTML")
    await state.set_state(SupportState.waiting_for_message)
    await callback.answer()


@dp.message(SupportState.waiting_for_message)
async def support_send_message(message: Message, state: FSMContext):
    if message.text and message.text.startswith("/"):
        await state.clear()
        return
    user = message.from_user
    await run_db(add_user_log, user.id, "support_message", f"Отправил сообщение: {(message.text or '')[:50]}")
    text = f"📩 <b>Сообщение от пользователя</b> @{user.username or 'без username'} (ID: {user.id})\n\n{escape_html(message.text or '')}"
    sent_to = 0
    for admin_id in ADMINS:
        try:
            await bot.send_message(admin_id, text, parse_mode="HTML")
            sent_to += 1
            if message.photo:
                file_id = message.photo[-1].file_id
                await bot.send_photo(admin_id, file_id, caption=f"Фото от @{user.username}")
            elif message.document:
                await bot.send_document(admin_id, message.document.file_id, caption=f"Документ от @{user.username}")
        except Exception as e:
            logging.warning(f"Не удалось отправить сообщение админу {admin_id}: {e}")
    if sent_to > 0:
        await message.answer("✅ Ваше сообщение отправлено автору. Он ответит вам здесь.",
                             reply_markup=back_to_main_keyboard(), parse_mode="HTML")
    else:
        await message.answer("❌ Не удалось отправить сообщение. Попробуйте позже.", parse_mode="HTML")
    await state.clear()


@dp.message(StateFilter(SupportState.waiting_for_message), F.text == "/cancel")
async def support_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Режим поддержки отменён.", reply_markup=main_menu_keyboard(), parse_mode="HTML")


# ===================== МОИ ЗАКАЗЫ =====================
@dp.callback_query(F.data == "my_orders")
async def cb_my_orders(callback: CallbackQuery):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "my_orders")
    await run_db(add_user_log, user_id, "my_orders", "Просмотрел свои заказы")
    orders = await run_db(get_user_orders, user_id)
    if not orders:
        text = "📋 У вас пока нет заказов."
        await update_message(callback, text, back_to_main_keyboard(), "HTML")
        await callback.answer()
        return
    text = "<b>📋 Ваши заказы:</b>\n\n"
    for order in orders:
        order_id, service, price, status, created_at, admin_price, order_code, rating, _, is_urgent = order
        status_text = {"pending": "⏳ Ожидает оплаты", "paid": "✅ Оплачен", "in_progress": "🔧 В работе",
                       "cancelled": "❌ Отменён"}.get(status, status)
        final_price = admin_price if admin_price > 0 else price
        created = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
        display_code = order_code or f"#{order_id}"
        urgent = "🔥 " if is_urgent else ""
        rating_str = f"⭐{rating}" if rating > 0 else ""
        text += f"• {urgent}{escape_html(display_code)}: {escape_html(service)} - {final_price} руб. ({status_text}) [{created}] {rating_str}\n"
    builder = InlineKeyboardBuilder()
    for order in orders:
        order_id, service, price, status, created_at, admin_price, order_code, rating, _, is_urgent = order
        display_code = order_code or f"#{order_id}"
        status_emoji = "✅" if status == "paid" else "⏳" if status == "pending" else "🔧" if status == "in_progress" else "❌"
        urgent = "🔥" if is_urgent else ""
        builder.button(text=f"{status_emoji}{urgent} {escape_html(display_code)} - {escape_html(service[:15])}",
                       callback_data=f"my_order_{order_id}")
    builder.button(text="🔙 Назад", callback_data="main_menu")
    builder.adjust(1)
    await update_message(callback, text, builder.as_markup(), "HTML")
    await callback.answer()


# ===================== ОСТАЛЬНЫЕ CALLBACK =====================
@dp.callback_query(F.data == "main_menu")
async def cb_main_menu(callback: CallbackQuery):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "main_menu")
    await run_db(add_user_log, user_id, "main_menu", "Вернулся в главное меню")
    text = "🏛️ <b>Sopranidi Corp.</b>\n\nВыберите необходимый сервис ниже 👇"
    if callback.message.photo:
        try:
            await callback.message.delete()
        except Exception:
            pass
        await callback.message.answer(text, reply_markup=main_menu_keyboard(), parse_mode="HTML")
    else:
        await update_message(callback, text, main_menu_keyboard(), "HTML")
    await callback.answer()


@dp.callback_query(F.data == "admin_menu")
async def cb_admin_menu(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Нет доступа")
        return
    stats = await run_db(get_stats)
    text = f"""
<b>🔐 Админ-панель Sopranidi Corp.</b>

👥 Пользователей: {stats['users']}
📦 Всего заказов: {stats['total_orders']}
✅ Оплаченных: {stats['paid_orders']}
⏳ Ожидают оплаты: {stats['pending_orders']}
🔧 В работе: {stats['in_progress']}
❌ Отменённых: {stats['cancelled_orders']}
💰 Доход: {stats['income']} руб.
⭐ Средняя оценка: {stats['avg_rating']}

📌 Диспетчер: {DISPATCHER_USERNAME}
👤 CEO: {CEO_USERNAME}
"""
    await update_message(callback, text, admin_menu_keyboard(callback.from_user.id), "HTML")
    await callback.answer()


# ===================== ПРИМЕРЫ РАБОТ =====================
@dp.callback_query(F.data == "examples")
async def cb_examples(callback: CallbackQuery):
    user_id = callback.from_user.id
    await run_db(update_user_action, user_id, "examples")
    await run_db(add_user_log, user_id, "examples", "Открыл примеры работ")
    builder = InlineKeyboardBuilder()
    builder.button(text="📄 Динамика цен на квартиры", callback_data="example_1")
    builder.button(text="💧 Экономия воды", callback_data="example_2")
    builder.button(text="🎱 План открытия бильярдной", callback_data="example_3")
    builder.button(text="🔙 Назад", callback_data="main_menu")
    builder.adjust(1)
    await update_message(callback, "📂 <b>Примеры выполненных работ</b>\n\nВыберите работу:", builder.as_markup(),
                         "HTML")
    await callback.answer()


@dp.callback_query(F.data.startswith("example_"))
async def send_example(callback: CallbackQuery):
    user_id = callback.from_user.id
    example_map = {
        "example_1": ("динамика цен на квартиры 2023-2026годов.pdf", "Динамика цен на квартиры 2023-2026"),
        "example_2": ("экономия воды.pdf", "Экономия воды"),
        "example_3": ("План открытия бильярдной.pdf", "План открытия бильярдной"),
    }
    file_name, title = example_map.get(callback.data, (None, None))
    if not file_name:
        await callback.answer("❌ Пример не найден", show_alert=True)
        return
    await run_db(add_user_log, user_id, "download_example", f"Скачал: {title}")
    try:
        file_path = EXAMPLES_DIR / file_name
        if not file_path.exists():
            if EXAMPLES_DIR.exists():
                pdf_files = [f for f in os.listdir(EXAMPLES_DIR) if f.endswith('.pdf')]
                if pdf_files:
                    file_path = EXAMPLES_DIR / pdf_files[0]
                    logging.info(f"Используем файл: {pdf_files[0]}")
                else:
                    await callback.answer("❌ Нет PDF-файлов", show_alert=True)
                    return
            else:
                await callback.answer("❌ Папка examples не найдена", show_alert=True)
                return
        file = FSInputFile(str(file_path))
        await callback.message.answer_document(document=file, caption=f"📄 {title}\n\n✅ Файл загружен!")
        await callback.answer("✅ Файл отправлен!")
    except Exception as e:
        logging.error(f"Ошибка отправки: {e}")
        await callback.answer("❌ Ошибка при отправке", show_alert=True)


# ===================== РАССЫЛКА (FSM) =====================
@dp.message(AdminBroadcastState.waiting_for_message)
async def broadcast_send(message: Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        await message.answer("⛔ Только супер-админ может делать рассылку.", parse_mode="HTML")
        await state.clear()
        return
    users = await run_db(get_all_users)
    sent = 0
    failed = 0
    for uid in users:
        try:
            await bot.send_message(uid[0], message.text, parse_mode="HTML")
            sent += 1
            await asyncio.sleep(0.1)
        except Exception as e:
            failed += 1
            logging.debug(f"Не удалось отправить рассылку {uid[0]}: {e}")
    await run_db(add_admin_log, message.from_user.id, "broadcast",
                 f"Отправил рассылку {sent} пользователям ({failed} неудачно)")
    await message.answer(f"✅ Рассылка выполнена. Отправлено {sent} пользователям, не доставлено {failed}.",
                         parse_mode="HTML")
    await state.clear()


# ===================== ФОНОВЫЕ ЗАДАЧИ =====================
async def birthday_checker_loop():
    while True:
        try:
            today_str = datetime.now().strftime("%d.%m")
            today_year = datetime.now().strftime("%Y")
            users = await run_db(get_users_with_birthday_today, today_str)
            for user_id, first_name, last_greet_year in users:
                if last_greet_year == today_year:
                    continue
                try:
                    await bot.send_message(user_id,
                                           f"🎉🎂 <b>С Днём Рождения, {escape_html(first_name or 'друг')}!</b>\n\nКоманда Sopranidi Corp. поздравляет вас и желает всего наилучшего! 🎁",
                                           parse_mode="HTML")
                    await run_db(mark_birthday_greeted, user_id, today_year)
                except Exception as e:
                    logging.warning(f"Не удалось поздравить пользователя {user_id}: {e}")
        except Exception as e:
            logging.error(f"Ошибка в birthday_checker_loop: {e}")
        await asyncio.sleep(3600)


async def poll_closer_loop():
    while True:
        try:
            expired = await run_db(get_expired_active_polls)
            for poll_id, question in expired:
                await run_db(close_poll, poll_id)
                logging.info(f"Голосование '{question}' автоматически закрыто (id={poll_id})")
        except Exception as e:
            logging.error(f"Ошибка в poll_closer_loop: {e}")
        await asyncio.sleep(300)


# ===================== ЗАПУСК БОТА =====================
async def main():
    init_db()
    logging.info("🚀 Бот Sopranidi Corp. запущен!")
    logging.info(f"📌 Диспетчер: {DISPATCHER_USERNAME}")
    logging.info(f"👤 CEO: {CEO_USERNAME}")
    logging.info(f"👥 Администраторы: {len(ADMINS)}")

    asyncio.create_task(birthday_checker_loop())
    asyncio.create_task(poll_closer_loop())

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())